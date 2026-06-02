"""XLSX import/export using openpyxl."""
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from engine import set_cell_input, recompute
from refs import make_ref, parse_ref
from values import Err, is_err

def export_xlsx(wb):
    out = openpyxl.Workbook()
    out.remove(out.active)
    for name in wb.sheet_order:
        sh = wb.sheets[name]
        ws = out.create_sheet(name)
        for (c, r), cell in sh.cells.items():
            xc = ws.cell(row=r+1, column=c+1)
            if cell.input is not None and isinstance(cell.input, str) and cell.input.startswith('='):
                xc.value = cell.input
            else:
                v = cell.value if cell.value is not None else cell.input
                if isinstance(v, Err):
                    xc.value = str(v)
                else:
                    xc.value = v
            if cell.format:
                xc.number_format = cell.format
            if cell.style:
                st = cell.style
                font_kw = {}
                if st.get('bold'): font_kw['bold'] = True
                if st.get('italic'): font_kw['italic'] = True
                if st.get('font_color'):
                    fc = st['font_color'].lstrip('#')
                    font_kw['color'] = 'FF' + fc.upper()
                if font_kw:
                    xc.font = Font(**font_kw)
                if st.get('bg_color'):
                    bg = st['bg_color'].lstrip('#')
                    xc.fill = PatternFill('solid', fgColor='FF' + bg.upper())
    # named ranges
    for nm, info in wb.names.items():
        try:
            rng = info['range']
            if '!' not in rng and info.get('sheet'):
                rng = info['sheet'] + '!' + rng
            elif '!' not in rng:
                rng = wb.sheet_order[0] + '!' + rng
            dn = DefinedName(name=nm, attr_text=rng)
            out.defined_names[nm] = dn
        except Exception: pass
    # Conditional formats
    try:
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Font as _F, PatternFill as _PF
        for cf in wb.cf_rules:
            if cf.get('sheet') not in out.sheetnames: continue
            ws = out[cf['sheet']]
            rule = cf['rule']
            style = rule.get('style', {})
            font_kw = {}
            if style.get('bold'): font_kw['bold'] = True
            if style.get('italic'): font_kw['italic'] = True
            if style.get('font_color'):
                font_kw['color'] = 'FF' + style['font_color'].lstrip('#').upper()
            font = _F(**font_kw) if font_kw else None
            fill = None
            if style.get('bg_color'):
                fill = _PF('solid', fgColor='FF' + style['bg_color'].lstrip('#').upper())
            op = rule.get('op')
            opmap = {'=':'equal','<>':'notEqual','<':'lessThan','>':'greaterThan','<=':'lessThanOrEqual','>=':'greaterThanOrEqual','between':'between'}
            xop = opmap.get(op, 'equal')
            val = rule.get('value')
            if op == 'between' and isinstance(val, list):
                formulas = [str(val[0]), str(val[1])]
            else:
                formulas = [str(val)]
            try:
                rrule = CellIsRule(operator=xop, formula=formulas, font=font, fill=fill)
                ws.conditional_formatting.add(cf['range'], rrule)
            except Exception: pass
    except Exception: pass
    # Data validations
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        for dv in wb.dv_rules:
            if dv.get('sheet') not in out.sheetnames: continue
            ws = out[dv['sheet']]
            rule = dv['rule']
            kind = rule.get('kind')
            if kind == 'list':
                vals = ','.join(rule.get('values', []))
                xdv = DataValidation(type='list', formula1='"' + vals + '"', allow_blank=True)
            elif kind == 'integer':
                xdv = DataValidation(type='whole', operator=rule.get('op','between'), formula1=str(rule.get('min','')), formula2=str(rule.get('max','')))
            elif kind == 'decimal':
                xdv = DataValidation(type='decimal', operator=rule.get('op','between'), formula1=str(rule.get('min','')), formula2=str(rule.get('max','')))
            elif kind == 'text_length':
                xdv = DataValidation(type='textLength', operator=rule.get('op','between'), formula1=str(rule.get('min','')), formula2=str(rule.get('max','')))
            else:
                continue
            xdv.add(dv['range'])
            ws.add_data_validation(xdv)
    except Exception: pass
    bio = io.BytesIO()
    out.save(bio)
    return bio.getvalue()

def import_xlsx(wb, data):
    inp = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    # clear existing sheets
    wb.sheets.clear(); wb.sheet_order.clear(); wb.deps.clear(); wb.precs.clear()
    for name in inp.sheetnames:
        wb.add_sheet(name)
        ws = inp[name]
        sh = wb.sheets[name]
        for row in ws.iter_rows():
            for xc in row:
                if xc.value is None: continue
                c = xc.column - 1; r = xc.row - 1
                v = xc.value
                if isinstance(v, str) and v.startswith('='):
                    try:
                        set_cell_input(wb, name, c, r, v)
                    except Exception:
                        cell = sh.get_or_create(c, r)
                        cell.input = v; cell.value = v; cell.kind = 'string'
                else:
                    if isinstance(v, bool):
                        s = 'TRUE' if v else 'FALSE'
                    elif isinstance(v, (int, float)):
                        s = str(v) if not (isinstance(v, float) and v.is_integer()) else str(int(v))
                    else:
                        s = str(v)
                    set_cell_input(wb, name, c, r, s)
                cell = sh.get_or_create(c, r)
                if xc.number_format and xc.number_format != 'General':
                    cell.format = xc.number_format
                style = {}
                try:
                    f = xc.font
                    if f and f.bold: style['bold'] = True
                    if f and f.italic: style['italic'] = True
                    if f and f.color and getattr(f.color, 'rgb', None) and isinstance(f.color.rgb, str) and len(f.color.rgb) == 8:
                        style['font_color'] = '#' + f.color.rgb[2:]
                    fill = xc.fill
                    if fill and getattr(fill, 'fgColor', None) and getattr(fill.fgColor, 'rgb', None):
                        rgb = fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) == 8 and rgb != '00000000':
                            style['bg_color'] = '#' + rgb[2:]
                except Exception: pass
                if style: cell.style = style
    # named ranges
    try:
        for dn in inp.defined_names:
            d = inp.defined_names[dn]
            wb.names[dn] = {'scope': 'workbook', 'sheet': None, 'range': d.attr_text}
    except Exception: pass
    # CF
    op_map = {'equal':'=','notEqual':'<>','lessThan':'<','greaterThan':'>','lessThanOrEqual':'<=','greaterThanOrEqual':'>=','between':'between'}
    for sname in inp.sheetnames:
        ws = inp[sname]
        try:
            for cf in list(ws.conditional_formatting):
                rng_str = str(cf.sqref) if hasattr(cf, 'sqref') else None
                for rule in ws.conditional_formatting[cf]:
                    op = op_map.get(getattr(rule, 'operator', None), '=')
                    formulas = list(rule.formula) if hasattr(rule, 'formula') and rule.formula else []
                    style = {}
                    if hasattr(rule, 'dxf') and rule.dxf:
                        try:
                            f = rule.dxf.font
                            if f and getattr(f, 'b', None): style['bold'] = True
                            if f and f.color and getattr(f.color, 'rgb', None) and isinstance(f.color.rgb, str) and len(f.color.rgb) == 8:
                                style['font_color'] = '#' + f.color.rgb[2:]
                            fill = rule.dxf.fill
                            if fill:
                                fg = getattr(fill, 'fgColor', None)
                                if fg and getattr(fg, 'rgb', None) and isinstance(fg.rgb, str) and len(fg.rgb) == 8 and fg.rgb != '00000000':
                                    style['bg_color'] = '#' + fg.rgb[2:]
                        except: pass
                    val = None
                    if formulas:
                        try: val = float(formulas[0])
                        except: val = formulas[0]
                        if op == 'between' and len(formulas) > 1:
                            try: val = [float(formulas[0]), float(formulas[1])]
                            except: val = formulas
                    wb.cf_rules.append({'sheet': sname, 'range': rng_str or 'A1', 'rule': {'kind':'cell_value','op':op,'value':val,'style':style}})
        except Exception: pass
        # DV
        try:
            for dv in ws.data_validations.dataValidation:
                t = dv.type
                rule = {}
                if t == 'list':
                    f1 = dv.formula1 or ''
                    rule = {'kind':'list', 'values': [s.strip().strip('"') for s in f1.split(',')]}
                elif t == 'whole':
                    rule = {'kind':'integer','op': op_map.get(dv.operator, 'between')}
                    try: rule['min'] = float(dv.formula1)
                    except: pass
                    try: rule['max'] = float(dv.formula2)
                    except: pass
                elif t == 'decimal':
                    rule = {'kind':'decimal','op': op_map.get(dv.operator, 'between')}
                    try: rule['min'] = float(dv.formula1)
                    except: pass
                    try: rule['max'] = float(dv.formula2)
                    except: pass
                elif t == 'textLength':
                    rule = {'kind':'text_length','op': op_map.get(dv.operator, 'between')}
                    try: rule['min'] = int(float(dv.formula1))
                    except: pass
                    try: rule['max'] = int(float(dv.formula2))
                    except: pass
                else: continue
                for r in dv.sqref.ranges if hasattr(dv.sqref, 'ranges') else [dv.sqref]:
                    wb.dv_rules.append({'sheet': sname, 'range': str(r), 'rule': rule})
        except Exception: pass
