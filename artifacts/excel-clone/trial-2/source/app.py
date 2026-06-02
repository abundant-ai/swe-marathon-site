from __future__ import annotations

import asyncio
import csv
import io
import json
import math
import os
import re
import statistics
import sys
import time
import zipfile
from collections import defaultdict, deque
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from itertools import product
from pathlib import Path
from typing import Any, Callable
from xml.etree import ElementTree as ET

import numpy as np
from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, Response
from openpyxl import Workbook as XLWorkbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from scipy import optimize, special, stats


DATA_DIR = Path("/app/data")
STATE_FILE = DATA_DIR / "state.json"
EXCEL_EPOCH = date(1899, 12, 30)
sys.setrecursionlimit(200000)


def now_ms() -> int:
    return int(time.time() * 1000)


class FormulaError(Exception):
    def __init__(self, code: str):
        super().__init__(code)
        self.code = code


class ApiError(Exception):
    def __init__(self, status: int, error: str, detail: str | None = None):
        self.status = status
        self.error = error
        self.detail = detail or error


ERR_DIV0 = "#DIV/0!"
ERR_NAME = "#NAME?"
ERR_NA = "#N/A"
ERR_NUM = "#NUM!"
ERR_SPILL = "#SPILL!"
ERR_CIRC = "#CIRC!"
ERR_VALUE = "#VALUE!"
ERRORS = {ERR_DIV0, ERR_NAME, ERR_NA, ERR_NUM, ERR_SPILL, ERR_CIRC, ERR_VALUE}


def is_error(v: Any) -> bool:
    return isinstance(v, str) and v in ERRORS


def col_to_num(col: str) -> int:
    n = 0
    for ch in col.upper().replace("$", ""):
        n = n * 26 + ord(ch) - 64
    return n


def num_to_col(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([0-9]{1,7})$")
CELL_FIND_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Za-z]{1,3}\$?[0-9]{1,7})(?![A-Za-z0-9_])")
SHEET_CELL_FIND_RE = re.compile(r"((?:'[^']+'|[A-Za-z_][A-Za-z0-9_ ]*)!)?(\$?[A-Za-z]{1,3}\$?[0-9]{1,7})(?::(\$?[A-Za-z]{1,3}\$?[0-9]{1,7}))?")


def split_ref(ref: str) -> tuple[int, int]:
    m = CELL_RE.match(ref)
    if not m:
        raise ApiError(400, "ParseError", f"bad cell ref {ref}")
    return int(m.group(2)), col_to_num(m.group(1))


def norm_ref(ref: str) -> str:
    r, c = split_ref(ref)
    return f"{num_to_col(c)}{r}"


def range_refs(range_text: str) -> list[str]:
    if ":" not in range_text:
        return [norm_ref(range_text)]
    a, b = range_text.split(":", 1)
    r1, c1 = split_ref(a)
    r2, c2 = split_ref(b)
    out = []
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            out.append(f"{num_to_col(c)}{r}")
    return out


def parse_sheet_ref(text: str, default_sheet: str) -> tuple[str, str]:
    if "!" in text:
        sheet, ref = text.split("!", 1)
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        return sheet, ref
    return default_sheet, text


def rect_bounds(range_text: str) -> tuple[int, int, int, int]:
    a, b = (range_text.split(":", 1) + [range_text])[:2] if ":" in range_text else (range_text, range_text)
    r1, c1 = split_ref(a)
    r2, c2 = split_ref(b)
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def ref_from_rc(row: int, col: int) -> str:
    return f"{num_to_col(col)}{row}"


def json_copy(v: Any) -> Any:
    return json.loads(json.dumps(v))


def default_workbook(wb_id: int, name: str) -> dict[str, Any]:
    return {
        "id": wb_id,
        "name": name,
        "created": now_ms(),
        "updated": now_ms(),
        "sheets": {"Sheet1": {"name": "Sheet1", "cells": {}}},
        "sheet_order": ["Sheet1"],
        "names": [],
        "conditional_formats": [],
        "validations": [],
        "settings": {
            "locale": "en-US",
            "iterative_calc": {"enabled": False, "max_iterations": 100, "max_change": 0.001},
        },
        "seq": 0,
        "events": [],
    }


class Store:
    def __init__(self):
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self.lock = asyncio.Lock()
        self.state = {"next_id": 1, "workbooks": {}}
        self.load()

    def load(self):
        if STATE_FILE.exists():
            try:
                self.state = json.loads(STATE_FILE.read_text())
                self.state.setdefault("next_id", 1)
                self.state.setdefault("workbooks", {})
            except Exception:
                self.state = {"next_id": 1, "workbooks": {}}

    def save(self):
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        tmp = STATE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(self.state, separators=(",", ":"), ensure_ascii=False))
        tmp.replace(STATE_FILE)

    def get(self, wb_id: int) -> dict[str, Any]:
        wb = self.state["workbooks"].get(str(wb_id))
        if not wb:
            raise ApiError(404, "NotFound")
        return wb


store = Store()
app = FastAPI()


@app.exception_handler(ApiError)
async def api_error_handler(_request: Request, exc: ApiError):
    return JSONResponse(status_code=exc.status, content={"error": exc.error, "detail": exc.detail})


def cell_kind(value: Any, explicit: str | None = None) -> str:
    if explicit:
        return explicit
    if value is None or value == "":
        return "empty"
    if is_error(value):
        return "error"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    return "string"


def coerce_literal(text: Any) -> tuple[Any, str]:
    if text is None:
        return None, "empty"
    if not isinstance(text, str):
        if isinstance(text, bool):
            return text, "bool"
        if isinstance(text, (int, float)):
            return text, "number"
        return str(text), "string"
    s = text
    if s == "":
        return "", "empty"
    if s.upper() == "TRUE":
        return True, "bool"
    if s.upper() == "FALSE":
        return False, "bool"
    try:
        if re.match(r"^[+-]?\d+$", s.strip()):
            return int(s), "number"
        if re.match(r"^[+-]?(?:\d+\.\d*|\.\d+|\d+\.)(?:[eE][+-]?\d+)?$", s.strip()) or re.match(r"^[+-]?\d+[eE][+-]?\d+$", s.strip()):
            return float(s), "number"
    except Exception:
        pass
    return s, "string"


def to_number(v: Any) -> float:
    if isinstance(v, RangeValue):
        vals = flatten(v.values())
        v = vals[0] if vals else 0
    if isinstance(v, list):
        vals = flatten(v)
        v = vals[0] if vals else 0
    if is_error(v):
        raise FormulaError(v)
    if v is None or v == "":
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        if not math.isfinite(float(v)):
            raise FormulaError(ERR_NUM)
        return float(v)
    try:
        return float(v)
    except Exception:
        raise FormulaError(ERR_VALUE)


def to_bool(v: Any) -> bool:
    if is_error(v):
        raise FormulaError(v)
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return v.upper() == "TRUE"
    return bool(v)


def scalar(v: Any) -> Any:
    if isinstance(v, RangeValue):
        vals = flatten(v.values())
        return vals[0] if vals else None
    if isinstance(v, list):
        vals = flatten(v)
        return vals[0] if vals else None
    return v


def flatten(v: Any) -> list[Any]:
    if isinstance(v, RangeValue):
        return flatten(v.values())
    if isinstance(v, list):
        out = []
        for x in v:
            if isinstance(x, (list, RangeValue)):
                out.extend(flatten(x))
            else:
                out.append(x)
        return out
    return [v]


def array2d(v: Any) -> list[list[Any]]:
    if isinstance(v, RangeValue):
        return v.values()
    if isinstance(v, list):
        if not v:
            return []
        if any(isinstance(x, list) for x in v):
            return [x if isinstance(x, list) else [x] for x in v]
        return [[x] for x in v]
    return [[v]]


def clean_number(x: float) -> int | float:
    if isinstance(x, bool):
        return x
    if abs(x - round(x)) < 1e-10:
        return int(round(x))
    return float(x)


@dataclass
class RangeValue:
    ctx: "EvalContext"
    sheet: str
    r1: int
    c1: int
    r2: int
    c2: int

    def refs(self) -> list[tuple[str, str]]:
        out = []
        for r in range(self.r1, self.r2 + 1):
            for c in range(self.c1, self.c2 + 1):
                out.append((self.sheet, ref_from_rc(r, c)))
        return out

    def values(self) -> list[list[Any]]:
        return [
            [self.ctx.get_cell_value(self.sheet, ref_from_rc(r, c)) for c in range(self.c1, self.c2 + 1)]
            for r in range(self.r1, self.r2 + 1)
        ]


@dataclass
class LambdaValue:
    params: list[str]
    body: Any
    env: dict[str, Any]
    ctx: "EvalContext"

    def __call__(self, args: list[Any]) -> Any:
        env = dict(self.env)
        for name, val in zip(self.params, args):
            env[name.upper()] = val
        return self.ctx.eval_node(self.body, env)


class Token:
    def __init__(self, typ: str, val: str):
        self.typ = typ
        self.val = val

    def __repr__(self):
        return f"Token({self.typ!r},{self.val!r})"


TOKEN_RE = re.compile(
    r"""\s*(?:
    (?P<str>"(?:[^"]|"")*")
    |(?P<num>(?:\d+\.\d*|\.\d+|\d+)(?:[eE][+-]?\d+)?)
    |(?P<op><>|<=|>=|[-+*/^&=<>%,():!])
    |(?P<name>\$?[A-Za-z_][A-Za-z0-9_.$]*|\$?[A-Za-z]{1,3}\$?\d{1,7})
    |(?P<sq>'[^']*(?:''[^']*)*')
    )""",
    re.X,
)


def tokenize(expr: str) -> list[Token]:
    out = []
    i = 0
    while i < len(expr):
        m = TOKEN_RE.match(expr, i)
        if not m:
            raise ApiError(400, "ParseError", f"unexpected token near {expr[i:i+10]}")
        i = m.end()
        typ = m.lastgroup or ""
        val = m.group(typ)
        if typ == "str":
            out.append(Token("str", val[1:-1].replace('""', '"')))
        elif typ == "num":
            out.append(Token("num", val))
        elif typ == "op":
            out.append(Token(val, val))
        elif typ == "sq":
            out.append(Token("name", val[1:-1].replace("''", "'")))
        else:
            out.append(Token("name", val))
    out.append(Token("EOF", ""))
    return out


class Parser:
    def __init__(self, text: str):
        self.tokens = tokenize(text)
        self.i = 0

    def peek(self) -> Token:
        return self.tokens[self.i]

    def pop(self, typ: str | None = None) -> Token:
        tok = self.peek()
        if typ and tok.typ != typ:
            raise ApiError(400, "ParseError", f"expected {typ}")
        self.i += 1
        return tok

    def parse(self):
        node = self.expr()
        if self.peek().typ != "EOF":
            raise ApiError(400, "ParseError", "trailing input")
        return node

    def expr(self):
        return self.compare()

    def compare(self):
        node = self.concat()
        while self.peek().typ in ("=", "<>", "<", ">", "<=", ">="):
            op = self.pop().typ
            rhs = self.concat()
            node = ("bin", op, node, rhs)
        return node

    def concat(self):
        node = self.add()
        while self.peek().typ == "&":
            self.pop()
            node = ("bin", "&", node, self.add())
        return node

    def add(self):
        node = self.mul()
        while self.peek().typ in ("+", "-"):
            op = self.pop().typ
            node = ("bin", op, node, self.mul())
        return node

    def mul(self):
        node = self.power()
        while self.peek().typ in ("*", "/"):
            op = self.pop().typ
            node = ("bin", op, node, self.power())
        return node

    def power(self):
        node = self.unary()
        if self.peek().typ == "^":
            self.pop()
            node = ("bin", "^", node, self.power())
        return node

    def unary(self):
        if self.peek().typ in ("+", "-"):
            op = self.pop().typ
            return ("unary", op, self.unary())
        return self.postfix()

    def postfix(self):
        node = self.primary()
        while True:
            if self.peek().typ == "%":
                self.pop()
                node = ("percent", node)
            elif self.peek().typ == "(":
                args = self.arglist()
                node = ("call_expr", node, args)
            else:
                return node

    def arglist(self):
        self.pop("(")
        args = []
        if self.peek().typ != ")":
            while True:
                args.append(self.expr())
                if self.peek().typ != ",":
                    break
                self.pop(",")
        self.pop(")")
        return args

    def primary(self):
        tok = self.peek()
        if tok.typ == "num":
            self.pop()
            return ("num", float(tok.val) if any(ch in tok.val for ch in ".eE") else int(tok.val))
        if tok.typ == "str":
            self.pop()
            return ("str", tok.val)
        if tok.typ == "(":
            self.pop()
            node = self.expr()
            self.pop(")")
            return node
        if tok.typ == "name":
            name = self.pop().val
            if self.peek().typ == "!":
                self.pop()
                ref_tok = self.pop("name")
                text = f"{name}!{ref_tok.val}"
                if self.peek().typ == ":":
                    self.pop()
                    end = self.pop("name").val
                    return ("range", text, end)
                return ("ref", text)
            if self.peek().typ == "(":
                args = self.arglist()
                return ("func", name.upper(), args)
            if CELL_RE.match(name):
                if self.peek().typ == ":":
                    self.pop()
                    end = self.pop("name").val
                    return ("range", name, end)
                return ("ref", name)
            return ("name", name.upper())
        raise ApiError(400, "ParseError", "expected expression")


def parse_formula(text: str):
    if not isinstance(text, str) or not text.startswith("="):
        raise ApiError(400, "ParseError")
    return Parser(text[1:]).parse()


def extract_deps(formula: str, sheet: str, wb: dict[str, Any]) -> set[tuple[str, str]]:
    deps: set[tuple[str, str]] = set()
    body = formula[1:] if formula.startswith("=") else formula
    for m in SHEET_CELL_FIND_RE.finditer(body):
        prefix, first, second = m.group(1), m.group(2), m.group(3)
        sh = sheet
        if prefix:
            sh = prefix[:-1]
            if sh.startswith("'") and sh.endswith("'"):
                sh = sh[1:-1].replace("''", "'")
        if second:
            r1, c1 = split_ref(first)
            r2, c2 = split_ref(second)
            for r in range(min(r1, r2), max(r1, r2) + 1):
                for c in range(min(c1, c2), max(c1, c2) + 1):
                    deps.add((sh, ref_from_rc(r, c)))
        else:
            try:
                deps.add((sh, norm_ref(first)))
            except ApiError:
                pass
    for n in wb.get("names", []):
        name = n.get("name", "").upper()
        if re.search(rf"(?<![A-Za-z0-9_]){re.escape(name)}(?![A-Za-z0-9_])", body, re.I):
            sh, rng = parse_sheet_ref(n["range"], sheet)
            for ref in range_refs(rng):
                deps.add((sh, ref))
    return deps


def translate_formula(text: str, locale: str) -> str:
    if not isinstance(text, str) or not text.startswith("="):
        return text
    mappings = {
        "de-DE": {"SUMME": "SUM", "MITTELWERT": "AVERAGE", "WENN": "IF", "WAHR": "TRUE", "FALSCH": "FALSE"},
        "fr-FR": {"SOMME": "SUM", "MOYENNE": "AVERAGE", "SI": "IF", "VRAI": "TRUE", "FAUX": "FALSE"},
        "es-ES": {"SUMA": "SUM", "PROMEDIO": "AVERAGE", "SI": "IF", "VERDADERO": "TRUE", "FALSO": "FALSE"},
    }
    if locale not in mappings:
        return text
    out = text
    for src, dst in mappings[locale].items():
        out = re.sub(rf"(?<![A-Za-z0-9_]){src}(?=\s*\(|\b)", dst, out, flags=re.I)
    # Convert decimal commas only when between digits, then argument separators.
    out = re.sub(r"(?<=\d),(?=\d)", ".", out)
    out = out.replace(";", ",")
    return out


class EvalContext:
    def __init__(self, wb: dict[str, Any], sheet: str, values: dict[tuple[str, str], Any], stack: set[tuple[str, str]] | None = None):
        self.wb = wb
        self.sheet = sheet
        self.values = values
        self.stack = stack or set()

    def get_cell_value(self, sheet: str, ref: str) -> Any:
        key = (sheet, norm_ref(ref))
        if key in self.values:
            return self.values[key]
        cell = self.wb["sheets"].get(sheet, {}).get("cells", {}).get(norm_ref(ref))
        if not cell:
            return None
        return cell.get("value")

    def eval_formula(self, formula: str) -> Any:
        try:
            return self.eval_node(parse_formula(formula), {})
        except ApiError:
            raise
        except FormulaError as e:
            return e.code
        except ZeroDivisionError:
            return ERR_DIV0
        except Exception:
            return ERR_VALUE

    def eval_node(self, node: Any, env: dict[str, Any]) -> Any:
        typ = node[0]
        if typ in ("num", "str"):
            return node[1]
        if typ == "ref":
            sh, ref = parse_sheet_ref(node[1], self.sheet)
            return self.get_cell_value(sh, ref)
        if typ == "range":
            start_text, end = node[1], node[2]
            sh, start = parse_sheet_ref(start_text, self.sheet)
            r1, c1 = split_ref(start)
            r2, c2 = split_ref(end)
            return RangeValue(self, sh, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        if typ == "name":
            name = node[1].upper()
            if name == "TRUE":
                return True
            if name == "FALSE":
                return False
            if name in env:
                return env[name]
            for n in self.wb.get("names", []):
                if n.get("name", "").upper() == name and (n.get("scope") == "workbook" or n.get("sheet") == self.sheet):
                    sh, rng = parse_sheet_ref(n["range"], self.sheet)
                    if ":" in rng:
                        r1, c1, r2, c2 = rect_bounds(rng)
                        return RangeValue(self, sh, r1, c1, r2, c2)
                    return self.get_cell_value(sh, rng)
            raise FormulaError(ERR_NAME)
        if typ == "unary":
            val = self.eval_node(node[2], env)
            return clean_number((1 if node[1] == "+" else -1) * to_number(val))
        if typ == "percent":
            return clean_number(to_number(self.eval_node(node[1], env)) / 100.0)
        if typ == "bin":
            op = node[1]
            if op in ("=", "<>", "<", ">", "<=", ">="):
                return self.compare_values(self.eval_node(node[2], env), self.eval_node(node[3], env), op)
            a = self.eval_node(node[2], env)
            b = self.eval_node(node[3], env)
            return self.binary(op, a, b)
        if typ == "func":
            return self.call_func(node[1], node[2], env)
        if typ == "call_expr":
            fn = self.eval_node(node[1], env)
            args = [self.eval_node(a, env) for a in node[2]]
            if isinstance(fn, LambdaValue):
                return fn(args)
            raise FormulaError(ERR_VALUE)
        raise FormulaError(ERR_VALUE)

    def compare_values(self, a: Any, b: Any, op: str) -> Any:
        if is_error(a):
            raise FormulaError(a)
        if is_error(b):
            raise FormulaError(b)
        aa, bb = array2d(a), array2d(b)
        if len(aa) > 1 or len(aa[0]) > 1 or len(bb) > 1 or len(bb[0]) > 1:
            rows = max(len(aa), len(bb))
            cols = max(max(len(r) for r in aa), max(len(r) for r in bb))
            return [[self.compare_values(aa[min(i, len(aa)-1)][min(j, len(aa[min(i, len(aa)-1)])-1)], bb[min(i, len(bb)-1)][min(j, len(bb[min(i, len(bb)-1)])-1)], op) for j in range(cols)] for i in range(rows)]
        a, b = scalar(a), scalar(b)
        if isinstance(a, str) or isinstance(b, str):
            av, bv = str(a).lower(), str(b).lower()
        else:
            av, bv = to_number(a), to_number(b)
        return {"=": av == bv, "<>": av != bv, "<": av < bv, ">": av > bv, "<=": av <= bv, ">=": av >= bv}[op]

    def binary(self, op: str, a: Any, b: Any) -> Any:
        if is_error(a):
            raise FormulaError(a)
        if is_error(b):
            raise FormulaError(b)
        if op == "&":
            return f"{format_plain(scalar(a))}{format_plain(scalar(b))}"
        aa, bb = array2d(a), array2d(b)
        if len(aa) > 1 or len(aa[0]) > 1 or len(bb) > 1 or len(bb[0]) > 1:
            rows = max(len(aa), len(bb))
            cols = max(max(len(r) for r in aa), max(len(r) for r in bb))
            out = []
            for i in range(rows):
                row = []
                for j in range(cols):
                    av = aa[min(i, len(aa)-1)][min(j, len(aa[min(i, len(aa)-1)])-1)]
                    bv = bb[min(i, len(bb)-1)][min(j, len(bb[min(i, len(bb)-1)])-1)]
                    row.append(self.binary(op, av, bv))
                out.append(row)
            return out
        x, y = to_number(a), to_number(b)
        if op == "+":
            return clean_number(x + y)
        if op == "-":
            return clean_number(x - y)
        if op == "*":
            return clean_number(x * y)
        if op == "/":
            if y == 0:
                raise FormulaError(ERR_DIV0)
            return clean_number(x / y)
        if op == "^":
            try:
                return clean_number(x ** y)
            except Exception:
                raise FormulaError(ERR_NUM)
        raise FormulaError(ERR_VALUE)

    def eval_args(self, arg_nodes: list[Any], env: dict[str, Any]) -> list[Any]:
        return [self.eval_node(a, env) for a in arg_nodes]

    def call_func(self, name: str, arg_nodes: list[Any], env: dict[str, Any]) -> Any:
        name = name.upper()
        if name == "LET":
            if len(arg_nodes) < 3 or len(arg_nodes) % 2 == 0:
                raise FormulaError(ERR_VALUE)
            local = dict(env)
            for i in range(0, len(arg_nodes) - 1, 2):
                n = arg_nodes[i]
                if n[0] != "name":
                    raise FormulaError(ERR_NAME)
                local[n[1].upper()] = self.eval_node(arg_nodes[i + 1], local)
            return self.eval_node(arg_nodes[-1], local)
        if name == "LAMBDA":
            if len(arg_nodes) < 2:
                raise FormulaError(ERR_VALUE)
            params = []
            for n in arg_nodes[:-1]:
                if n[0] != "name":
                    raise FormulaError(ERR_NAME)
                params.append(n[1].upper())
            return LambdaValue(params, arg_nodes[-1], dict(env), self)
        if name == "IF":
            cond = self.eval_node(arg_nodes[0], env) if arg_nodes else False
            branch = 1 if to_bool(cond) else 2
            if branch < len(arg_nodes):
                return self.eval_node(arg_nodes[branch], env)
            return False
        if name == "IFERROR":
            try:
                val = self.eval_node(arg_nodes[0], env)
                if is_error(val):
                    return self.eval_node(arg_nodes[1], env)
                return val
            except FormulaError:
                return self.eval_node(arg_nodes[1], env)
        if name in ("ISERROR", "ISERR", "ISNA"):
            try:
                val = self.eval_node(arg_nodes[0], env)
                if name == "ISNA":
                    return val == ERR_NA
                if name == "ISERR":
                    return is_error(val) and val != ERR_NA
                return is_error(val)
            except FormulaError as e:
                if name == "ISNA":
                    return e.code == ERR_NA
                if name == "ISERR":
                    return e.code != ERR_NA
                return True
        if name in env and isinstance(env[name], LambdaValue):
            return env[name]([self.eval_node(a, env) for a in arg_nodes])
        if name == "OFFSET":
            if not arg_nodes:
                raise FormulaError(ERR_VALUE)
            ref = self.ref_node_value(arg_nodes[0], env)
            args = [self.eval_node(a, env) for a in arg_nodes[1:]]
            rows = int(to_number(args[0] if len(args) > 0 else 0))
            cols = int(to_number(args[1] if len(args) > 1 else 0))
            height = int(to_number(args[2] if len(args) > 2 else (ref.r2 - ref.r1 + 1)))
            width = int(to_number(args[3] if len(args) > 3 else (ref.c2 - ref.c1 + 1)))
            return RangeValue(self, ref.sheet, ref.r1 + rows, ref.c1 + cols, ref.r1 + rows + height - 1, ref.c1 + cols + width - 1)
        if name == "INDIRECT":
            text = str(self.eval_node(arg_nodes[0], env)).replace("$", "")
            sh, rng = parse_sheet_ref(text, self.sheet)
            if ":" in rng:
                r1, c1, r2, c2 = rect_bounds(rng)
                return RangeValue(self, sh, r1, c1, r2, c2)
            return self.get_cell_value(sh, rng)
        if name == "IFNA":
            try:
                val = self.eval_node(arg_nodes[0], env)
                return self.eval_node(arg_nodes[1], env) if val == ERR_NA else val
            except FormulaError as e:
                return self.eval_node(arg_nodes[1], env) if e.code == ERR_NA else e.code
        if name in ("ROW", "COLUMN", "ROWS", "COLUMNS"):
            ref = self.ref_node_value(arg_nodes[0], env) if arg_nodes else RangeValue(self, self.sheet, 1, 1, 1, 1)
            if name == "ROW":
                return ref.r1
            if name == "COLUMN":
                return ref.c1
            if name == "ROWS":
                return ref.r2 - ref.r1 + 1
            return ref.c2 - ref.c1 + 1
        args = self.eval_args(arg_nodes, env)
        fn = FUNCTIONS.get(name)
        if fn:
            return fn(self, *args)
        raise FormulaError(ERR_NAME)

    def ref_node_value(self, node: Any, env: dict[str, Any]) -> RangeValue:
        if node[0] == "ref":
            sh, ref = parse_sheet_ref(node[1], self.sheet)
            r, c = split_ref(ref)
            return RangeValue(self, sh, r, c, r, c)
        if node[0] == "range":
            start_text, end = node[1], node[2]
            sh, start = parse_sheet_ref(start_text, self.sheet)
            r1, c1 = split_ref(start)
            r2, c2 = split_ref(end)
            return RangeValue(self, sh, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))
        val = self.eval_node(node, env)
        if isinstance(val, RangeValue):
            return val
        raise FormulaError(ERR_VALUE)


def format_plain(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and abs(v - round(v)) < 1e-10:
        return str(int(round(v)))
    return str(v)


def numbers(*args: Any, ignore_text: bool = True) -> list[float]:
    out = []
    for v in args:
        for x in flatten(v):
            if is_error(x):
                raise FormulaError(x)
            if isinstance(x, bool):
                out.append(1.0 if x else 0.0)
            elif isinstance(x, (int, float)) and not isinstance(x, bool):
                out.append(float(x))
            elif not ignore_text and x not in (None, ""):
                out.append(to_number(x))
    return out


def criteria_match(value: Any, crit: Any) -> bool:
    if isinstance(crit, (int, float)):
        try:
            return to_number(value) == float(crit)
        except FormulaError:
            return False
    s = str(crit)
    m = re.match(r"^(<=|>=|<>|=|<|>)(.*)$", s)
    if m:
        op, rhs = m.group(1), m.group(2)
        try:
            lhs_num = to_number(value)
            rhs_num = float(rhs)
            return {"=": lhs_num == rhs_num, "<>": lhs_num != rhs_num, "<": lhs_num < rhs_num, ">": lhs_num > rhs_num, "<=": lhs_num <= rhs_num, ">=": lhs_num >= rhs_num}[op]
        except Exception:
            lhs, rhs_s = str(value).lower(), rhs.lower()
            return {"=": lhs == rhs_s, "<>": lhs != rhs_s, "<": lhs < rhs_s, ">": lhs > rhs_s, "<=": lhs <= rhs_s, ">=": lhs >= rhs_s}[op]
    if "*" in s or "?" in s:
        pat = "^" + re.escape(s).replace(r"\*", ".*").replace(r"\?", ".") + "$"
        return re.match(pat, str(value), re.I) is not None
    return str(value).lower() == s.lower()


def func_sum(ctx, *args):
    return clean_number(sum(numbers(*args)))


def func_average(ctx, *args):
    ns = numbers(*args)
    return clean_number(sum(ns) / len(ns)) if ns else ERR_DIV0


def func_min(ctx, *args):
    ns = numbers(*args)
    return clean_number(min(ns)) if ns else 0


def func_max(ctx, *args):
    ns = numbers(*args)
    return clean_number(max(ns)) if ns else 0


def func_count(ctx, *args):
    return len(numbers(*args))


def func_product(ctx, *args):
    p = 1.0
    for n in numbers(*args):
        p *= n
    return clean_number(p)


def func_countif(ctx, rng, crit):
    return sum(1 for v in flatten(rng) if criteria_match(v, crit))


def func_sumif(ctx, rng, crit, sum_rng=None):
    vals = flatten(rng)
    sums = flatten(sum_rng if sum_rng is not None else rng)
    total = 0.0
    for i, v in enumerate(vals):
        if criteria_match(v, crit) and i < len(sums):
            try:
                total += to_number(sums[i])
            except FormulaError:
                pass
    return clean_number(total)


def func_averageif(ctx, rng, crit, avg_rng=None):
    vals = flatten(rng)
    avs = flatten(avg_rng if avg_rng is not None else rng)
    ns = []
    for i, v in enumerate(vals):
        if criteria_match(v, crit) and i < len(avs):
            try:
                ns.append(to_number(avs[i]))
            except FormulaError:
                pass
    return clean_number(sum(ns) / len(ns)) if ns else ERR_DIV0


def _multi_criteria(args):
    base = flatten(args[0])
    pairs = [(flatten(args[i]), args[i + 1]) for i in range(1, len(args), 2)]
    idxs = []
    for i in range(len(base)):
        if all(i < len(rng) and criteria_match(rng[i], crit) for rng, crit in pairs):
            idxs.append(i)
    return base, idxs


def func_sumifs(ctx, sum_range, *args):
    base, idxs = _multi_criteria((sum_range,) + args)
    return clean_number(sum(to_number(base[i]) for i in idxs))


def func_countifs(ctx, *args):
    base = flatten(args[0])
    pairs = [(flatten(args[i]), args[i + 1]) for i in range(0, len(args), 2)]
    return sum(1 for i in range(len(base)) if all(i < len(rng) and criteria_match(rng[i], crit) for rng, crit in pairs))


def func_averageifs(ctx, avg_range, *args):
    base, idxs = _multi_criteria((avg_range,) + args)
    ns = [to_number(base[i]) for i in idxs]
    return clean_number(sum(ns) / len(ns)) if ns else ERR_DIV0


def func_sequence(ctx, rows, cols=1, start=1, step=1):
    rows, cols = int(to_number(rows)), int(to_number(cols))
    start, step = to_number(start), to_number(step)
    return [[clean_number(start + (r * cols + c) * step) for c in range(cols)] for r in range(rows)]


def func_map(ctx, arr, lam):
    if not isinstance(lam, LambdaValue):
        raise FormulaError(ERR_VALUE)
    data = array2d(arr)
    return [[lam([v]) for v in row] for row in data]


def func_reduce(ctx, initial, arr, lam):
    if not isinstance(lam, LambdaValue):
        raise FormulaError(ERR_VALUE)
    acc = initial
    for v in flatten(arr):
        acc = lam([acc, v])
    return acc


def func_filter(ctx, arr, include, if_empty=None):
    data = array2d(arr)
    inc = array2d(include)
    out = []
    for i, row in enumerate(data):
        keep = False
        if len(inc) == len(data) and len(inc[i]) == 1:
            keep = to_bool(inc[i][0])
        elif len(inc) == 1 and len(inc[0]) == len(row):
            keep = any(to_bool(x) for x in inc[0])
        elif i < len(inc):
            keep = any(to_bool(x) for x in inc[i])
        if keep:
            out.append(row)
    if not out:
        return if_empty if if_empty is not None else ERR_NA
    return out


def func_sort(ctx, arr, sort_index=1, sort_order=1, by_col=False):
    data = array2d(arr)
    idx = int(to_number(sort_index)) - 1
    rev = to_number(sort_order) < 0
    return sorted(data, key=lambda r: (r[idx] if idx < len(r) else None), reverse=rev)


def func_unique(ctx, arr):
    seen = set()
    out = []
    for row in array2d(arr):
        key = tuple(row)
        if key not in seen:
            seen.add(key)
            out.append(row)
    return out


def func_byrow(ctx, arr, lam):
    if not isinstance(lam, LambdaValue):
        raise FormulaError(ERR_VALUE)
    return [[lam([row])] for row in array2d(arr)]


def func_bycol(ctx, arr, lam):
    if not isinstance(lam, LambdaValue):
        raise FormulaError(ERR_VALUE)
    data = array2d(arr)
    cols = max(len(r) for r in data) if data else 0
    return [[lam([[row[c] for row in data]]) for c in range(cols)]]


def func_vlookup(ctx, lookup, table, col_index, approx=False):
    lookup = scalar(lookup)
    data = array2d(table)
    ci = int(to_number(col_index)) - 1
    for row in data:
        if row and row[0] == lookup:
            return row[ci] if ci < len(row) else ERR_NA
    if to_bool(approx):
        best = None
        for row in data:
            if not row:
                continue
            try:
                if to_number(row[0]) <= to_number(lookup):
                    best = row
            except Exception:
                if str(row[0]) <= str(lookup):
                    best = row
        if best is not None:
            return best[ci] if ci < len(best) else ERR_NA
    return ERR_NA


def func_hlookup(ctx, lookup, table, row_index, approx=False):
    lookup = scalar(lookup)
    data = array2d(table)
    ri = int(to_number(row_index)) - 1
    if not data:
        return ERR_NA
    for j, v in enumerate(data[0]):
        if v == lookup:
            return data[ri][j] if ri < len(data) and j < len(data[ri]) else ERR_NA
    if to_bool(approx):
        best = None
        for j, v in enumerate(data[0]):
            try:
                if to_number(v) <= to_number(lookup):
                    best = j
            except Exception:
                if str(v) <= str(lookup):
                    best = j
        if best is not None:
            return data[ri][best] if ri < len(data) and best < len(data[ri]) else ERR_NA
    return ERR_NA


def func_match(ctx, lookup, arr, match_type=0):
    vals = flatten(arr)
    mt = int(to_number(match_type))
    for i, v in enumerate(vals, 1):
        if v == lookup or (mt == 0 and isinstance(lookup, str) and ("*" in lookup or "?" in lookup) and criteria_match(v, lookup)):
            return i
    if mt:
        best_i = None
        for i, v in enumerate(vals, 1):
            try:
                ok = to_number(v) <= to_number(lookup) if mt > 0 else to_number(v) >= to_number(lookup)
            except Exception:
                ok = str(v) <= str(lookup) if mt > 0 else str(v) >= str(lookup)
            if ok:
                best_i = i
        if best_i is not None:
            return best_i
    return ERR_NA


def func_index(ctx, arr, row_num=1, col_num=1):
    data = array2d(arr)
    r = int(to_number(row_num)) - 1
    c = int(to_number(col_num)) - 1
    try:
        return data[r][c]
    except Exception:
        return ERR_NA


def func_xlookup(ctx, lookup, lookup_arr, return_arr, if_not_found=ERR_NA, match_mode=0, search_mode=1):
    keys = flatten(lookup_arr)
    vals = flatten(return_arr)
    order = range(len(keys) - 1, -1, -1) if int(to_number(search_mode)) < 0 else range(len(keys))
    mm = int(to_number(match_mode))
    for i in order:
        k = keys[i]
        if k == lookup or (mm == 2 and isinstance(lookup, str) and criteria_match(k, lookup)):
            return vals[i] if i < len(vals) else ERR_NA
    if mm in (-1, 1):
        best = None
        for i, k in enumerate(keys):
            try:
                ok = to_number(k) <= to_number(lookup) if mm == -1 else to_number(k) >= to_number(lookup)
            except Exception:
                ok = str(k) <= str(lookup) if mm == -1 else str(k) >= str(lookup)
            if ok:
                best = i
                if mm == 1:
                    break
        if best is not None:
            return vals[best] if best < len(vals) else ERR_NA
    return if_not_found


def excel_date(y, m, d):
    y, m, d = int(to_number(y)), int(to_number(m)), int(to_number(d))
    month0 = m - 1
    yy = y + month0 // 12
    mm = month0 % 12 + 1
    base = date(yy, mm, 1) + timedelta(days=d - 1)
    return (base - EXCEL_EPOCH).days


def serial_to_date(n) -> date:
    return EXCEL_EPOCH + timedelta(days=int(to_number(n)))


def excel_time(h=0, m=0, s=0):
    return clean_number((to_number(h) * 3600 + to_number(m) * 60 + to_number(s)) / 86400)


def eomonth(serial, months):
    dt = serial_to_date(serial)
    month0 = dt.month - 1 + int(to_number(months)) + 1
    yy = dt.year + month0 // 12
    mm = month0 % 12 + 1
    first_next = date(yy, mm, 1)
    return (first_next - timedelta(days=1) - EXCEL_EPOCH).days


def edate(serial, months):
    dt = serial_to_date(serial)
    month0 = dt.month - 1 + int(to_number(months))
    yy = dt.year + month0 // 12
    mm = month0 % 12 + 1
    last = (date(yy + (mm // 12), (mm % 12) + 1, 1) - timedelta(days=1)).day if mm < 12 else 31
    return (date(yy, mm, min(dt.day, last)) - EXCEL_EPOCH).days


def networkdays(start, end, holidays=None):
    s, e = serial_to_date(start), serial_to_date(end)
    if e < s:
        s, e = e, s
    h = {serial_to_date(x) for x in flatten(holidays)} if holidays is not None else set()
    return sum(1 for i in range((e - s).days + 1) if (s + timedelta(days=i)).weekday() < 5 and (s + timedelta(days=i)) not in h)


def workday(start, days, holidays=None):
    d = serial_to_date(start)
    n = int(to_number(days))
    h = {serial_to_date(x) for x in flatten(holidays)} if holidays is not None else set()
    step = 1 if n >= 0 else -1
    left = abs(n)
    while left:
        d += timedelta(days=step)
        if d.weekday() < 5 and d not in h:
            left -= 1
    return (d - EXCEL_EPOCH).days


def textsplit(ctx, text, col_delim, row_delim=None):
    rows = str(text).split(str(row_delim)) if row_delim is not None else [str(text)]
    return [r.split(str(col_delim)) for r in rows]


def textbefore(ctx, text, delim, instance_num=1):
    parts = str(text).split(str(delim))
    n = int(to_number(instance_num))
    return str(delim).join(parts[:n]) if len(parts) > n else ERR_NA


def textafter(ctx, text, delim, instance_num=1):
    parts = str(text).split(str(delim))
    n = int(to_number(instance_num))
    return str(delim).join(parts[n:]) if len(parts) > n else ERR_NA


def pmt(rate, nper, pv, fv=0, typ=0):
    rate, nper, pv, fv, typ = map(to_number, (rate, nper, pv, fv, typ))
    if rate == 0:
        return clean_number(-(pv + fv) / nper)
    return clean_number(-(rate * (fv + pv * (1 + rate) ** nper)) / ((1 + rate * typ) * ((1 + rate) ** nper - 1)))


def excel_round(value, digits=0):
    x = to_number(value)
    n = int(to_number(digits))
    factor = 10 ** n
    if x >= 0:
        return clean_number(math.floor(x * factor + 0.5) / factor)
    return clean_number(math.ceil(x * factor - 0.5) / factor)


def roundup(value, digits=0):
    x = to_number(value)
    n = int(to_number(digits))
    factor = 10 ** n
    return clean_number((math.ceil(abs(x) * factor) / factor) * (1 if x >= 0 else -1))


def rounddown(value, digits=0):
    x = to_number(value)
    n = int(to_number(digits))
    factor = 10 ** n
    return clean_number((math.floor(abs(x) * factor) / factor) * (1 if x >= 0 else -1))


def sumproduct(*args):
    arrays = [numbers(a) for a in args]
    if not arrays:
        return 0
    n = min(len(a) for a in arrays)
    total = 0.0
    for i in range(n):
        prod_v = 1.0
        for arr in arrays:
            prod_v *= arr[i]
        total += prod_v
    return clean_number(total)


def textjoin(delim, ignore_empty, *args):
    vals = [format_plain(v) for v in flatten(list(args))]
    if to_bool(ignore_empty):
        vals = [v for v in vals if v != ""]
    return str(delim).join(vals)


def excel_text(value, fmt):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return format_display(value, str(fmt))
    return format_plain(value)


def choose(index, *values):
    i = int(to_number(index)) - 1
    if i < 0 or i >= len(values):
        return ERR_VALUE
    return values[i]


def pv(rate, nper, pmtv, fv=0, typ=0):
    rate, nper, pmtv, fv, typ = map(to_number, (rate, nper, pmtv, fv, typ))
    if rate == 0:
        return clean_number(-fv - pmtv * nper)
    return clean_number(-(fv + pmtv * (1 + rate * typ) * ((1 + rate) ** nper - 1) / rate) / ((1 + rate) ** nper))


def fv(rate, nper, pmtv, pv0=0, typ=0):
    rate, nper, pmtv, pv0, typ = map(to_number, (rate, nper, pmtv, pv0, typ))
    if rate == 0:
        return clean_number(-(pv0 + pmtv * nper))
    return clean_number(-(pv0 * (1 + rate) ** nper + pmtv * (1 + rate * typ) * ((1 + rate) ** nper - 1) / rate))


def npv(rate, *vals):
    r = to_number(rate)
    return clean_number(sum(to_number(v) / ((1 + r) ** i) for i, v in enumerate(flatten(list(vals)), 1)))


def irr(values, guess=0.1):
    vals = [to_number(v) for v in flatten(values)]
    try:
        return clean_number(optimize.newton(lambda r: sum(v / ((1 + r) ** i) for i, v in enumerate(vals)), to_number(guess)))
    except Exception:
        return ERR_NUM


def nper(rate, pmtv, pv0, fv0=0, typ=0):
    rate, pmtv, pv0, fv0, typ = map(to_number, (rate, pmtv, pv0, fv0, typ))
    if rate == 0:
        return clean_number(-(pv0 + fv0) / pmtv)
    try:
        return clean_number(math.log((pmtv * (1 + rate * typ) - fv0 * rate) / (pv0 * rate + pmtv * (1 + rate * typ))) / math.log(1 + rate))
    except Exception:
        return ERR_NUM


def rank(ctx, number, ref, order=0):
    vals = sorted(numbers(ref), reverse=(to_number(order) == 0))
    num = to_number(number)
    return vals.index(num) + 1 if num in vals else ERR_NA


def database_rows(database):
    data = array2d(database)
    if not data:
        return [], []
    return data[0], data[1:]


def db_field_index(headers, field):
    if isinstance(field, (int, float)):
        return int(field) - 1
    low = str(field).lower()
    for i, h in enumerate(headers):
        if str(h).lower() == low:
            return i
    return -1


def db_matches(row, headers, criteria):
    crit = array2d(criteria)
    if len(crit) < 2:
        return True
    crit_headers = crit[0]
    for crow in crit[1:]:
        ok = True
        for i, c in enumerate(crow):
            if c in (None, ""):
                continue
            idx = db_field_index(headers, crit_headers[i])
            if idx < 0 or idx >= len(row) or not criteria_match(row[idx], c):
                ok = False
                break
        if ok:
            return True
    return False


def db_values(database, field, criteria):
    headers, rows = database_rows(database)
    idx = db_field_index(headers, field)
    return [row[idx] for row in rows if idx >= 0 and idx < len(row) and db_matches(row, headers, criteria)]


FUNCTIONS: dict[str, Callable[..., Any]] = {
    "SUM": func_sum, "AVERAGE": func_average, "MIN": func_min, "MAX": func_max, "COUNT": func_count, "PRODUCT": func_product,
    "COUNTA": lambda c, *a: len([x for x in flatten(list(a)) if x not in (None, "")]),
    "COUNTBLANK": lambda c, a: len([x for x in flatten(a) if x in (None, "")]),
    "SUMPRODUCT": lambda c, *a: sumproduct(*a),
    "COUNTIF": func_countif, "SUMIF": func_sumif, "AVERAGEIF": func_averageif, "SUMIFS": func_sumifs, "COUNTIFS": func_countifs, "AVERAGEIFS": func_averageifs,
    "POWER": lambda c, x, y: clean_number(to_number(x) ** to_number(y)),
    "SQRT": lambda c, x: clean_number(math.sqrt(to_number(x))) if to_number(x) >= 0 else ERR_NUM,
    "SQRTPI": lambda c, x: clean_number(math.sqrt(to_number(x) * math.pi)) if to_number(x) >= 0 else ERR_NUM,
    "MOD": lambda c, x, y: ERR_DIV0 if to_number(y) == 0 else clean_number(to_number(x) % to_number(y)),
    "ABS": lambda c, x: clean_number(abs(to_number(x))),
    "ROUND": lambda c, x, n=0: excel_round(x, n),
    "ROUNDUP": lambda c, x, n=0: roundup(x, n),
    "ROUNDDOWN": lambda c, x, n=0: rounddown(x, n),
    "MROUND": lambda c, x, m: clean_number(round(to_number(x) / to_number(m)) * to_number(m)),
    "CEILING": lambda c, x, sig=1: clean_number(math.ceil(to_number(x) / to_number(sig)) * to_number(sig)),
    "FLOOR": lambda c, x, sig=1: clean_number(math.floor(to_number(x) / to_number(sig)) * to_number(sig)),
    "INT": lambda c, x: math.floor(to_number(x)),
    "PI": lambda c: math.pi,
    "SIN": lambda c, x: math.sin(to_number(x)), "COS": lambda c, x: math.cos(to_number(x)), "TAN": lambda c, x: math.tan(to_number(x)),
    "ASIN": lambda c, x: math.asin(to_number(x)), "ACOS": lambda c, x: math.acos(to_number(x)), "ATAN": lambda c, x: math.atan(to_number(x)), "ATAN2": lambda c, x, y: math.atan2(to_number(x), to_number(y)),
    "SINH": lambda c, x: math.sinh(to_number(x)), "COSH": lambda c, x: math.cosh(to_number(x)), "TANH": lambda c, x: math.tanh(to_number(x)),
    "RADIANS": lambda c, x: math.radians(to_number(x)), "DEGREES": lambda c, x: math.degrees(to_number(x)),
    "FACT": lambda c, x: math.factorial(int(to_number(x))) if to_number(x) >= 0 else ERR_NUM,
    "COMBIN": lambda c, n, k: math.comb(int(to_number(n)), int(to_number(k))), "PERMUT": lambda c, n, k: math.perm(int(to_number(n)), int(to_number(k))),
    "GCD": lambda c, *a: math.gcd(*[int(abs(x)) for x in numbers(*a)]), "LCM": lambda c, *a: math.lcm(*[int(abs(x)) for x in numbers(*a)]),
    "LOG": lambda c, x, b=10: clean_number(math.log(to_number(x), to_number(b))) if to_number(x) > 0 and to_number(b) > 0 else ERR_NUM,
    "LN": lambda c, x: clean_number(math.log(to_number(x))) if to_number(x) > 0 else ERR_NUM,
    "LOG10": lambda c, x: clean_number(math.log10(to_number(x))) if to_number(x) > 0 else ERR_NUM,
    "EXP": lambda c, x: clean_number(math.exp(to_number(x))),
    "SIGN": lambda c, x: 1 if to_number(x) > 0 else (-1 if to_number(x) < 0 else 0),
    "MEDIAN": lambda c, *a: clean_number(statistics.median(numbers(*a))),
    "STDEV": lambda c, *a: clean_number(statistics.stdev(numbers(*a))), "STDEVP": lambda c, *a: clean_number(statistics.pstdev(numbers(*a))),
    "VAR": lambda c, *a: clean_number(statistics.variance(numbers(*a))), "VARP": lambda c, *a: clean_number(statistics.pvariance(numbers(*a))),
    "PERCENTILE": lambda c, a, k: clean_number(float(np.percentile(numbers(a), to_number(k) * 100))),
    "QUARTILE": lambda c, a, q: clean_number(float(np.percentile(numbers(a), to_number(q) * 25))),
    "RANK": rank,
    "CORREL": lambda c, a, b: clean_number(float(np.corrcoef(numbers(a), numbers(b))[0, 1])),
    "COVAR": lambda c, a, b: clean_number(float(np.cov(numbers(a), numbers(b), bias=True)[0, 1])),
    "SLOPE": lambda c, y, x: clean_number(float(np.polyfit(numbers(x), numbers(y), 1)[0])),
    "INTERCEPT": lambda c, y, x: clean_number(float(np.polyfit(numbers(x), numbers(y), 1)[1])),
    "FORECAST": lambda c, x, y, known_x: clean_number(float(np.polyval(np.polyfit(numbers(known_x), numbers(y), 1), to_number(x)))),
    "NORM.DIST": lambda c, x, mean, sd, cum=True: float(stats.norm.cdf(to_number(x), to_number(mean), to_number(sd)) if to_bool(cum) else stats.norm.pdf(to_number(x), to_number(mean), to_number(sd))),
    "NORM.INV": lambda c, p, mean, sd: float(stats.norm.ppf(to_number(p), to_number(mean), to_number(sd))),
    "NORM.S.DIST": lambda c, z, cum=True: float(stats.norm.cdf(to_number(z)) if to_bool(cum) else stats.norm.pdf(to_number(z))),
    "BINOM.DIST": lambda c, x, n, p, cum=True: float(stats.binom.cdf(int(to_number(x)), int(to_number(n)), to_number(p)) if to_bool(cum) else stats.binom.pmf(int(to_number(x)), int(to_number(n)), to_number(p))),
    "POISSON.DIST": lambda c, x, mean, cum=True: float(stats.poisson.cdf(int(to_number(x)), to_number(mean)) if to_bool(cum) else stats.poisson.pmf(int(to_number(x)), to_number(mean))),
    "EXPON.DIST": lambda c, x, lam, cum=True: float(stats.expon.cdf(to_number(x), scale=1 / to_number(lam)) if to_bool(cum) else lam * math.exp(-to_number(lam) * to_number(x))),
    "GAMMA": lambda c, x: float(special.gamma(to_number(x))), "GAMMALN": lambda c, x: float(special.gammaln(to_number(x))),
    "PMT": lambda c, *a: pmt(*a), "FV": lambda c, *a: fv(*a), "PV": lambda c, *a: pv(*a), "NPV": lambda c, *a: npv(*a), "IRR": lambda c, *a: irr(*a),
    "RATE": lambda c, nper0, pmt0, pv0, fv0=0, typ=0, guess=0.1: clean_number(optimize.newton(lambda r: fv(r, nper0, pmt0, pv0, typ) + to_number(fv0), to_number(guess))),
    "NPER": lambda c, *a: nper(*a), "SLN": lambda c, cost, salvage, life: clean_number((to_number(cost) - to_number(salvage)) / to_number(life)),
    "CUMIPMT": lambda c, rate, nper0, pv0, start, end, typ=0: clean_number(sum((pv(rate, i - 1, pmt(rate, nper0, pv0), pv0) * to_number(rate)) for i in range(int(to_number(start)), int(to_number(end)) + 1))),
    "DATE": lambda c, y, m, d: excel_date(y, m, d), "YEAR": lambda c, x: serial_to_date(x).year, "MONTH": lambda c, x: serial_to_date(x).month, "DAY": lambda c, x: serial_to_date(x).day,
    "WEEKDAY": lambda c, x, typ=1: ((serial_to_date(x).weekday() + 1) % 7) + 1 if int(to_number(typ)) == 1 else serial_to_date(x).weekday() + 1,
    "DAYS": lambda c, e, s: int(to_number(e) - to_number(s)), "TIME": lambda c, h=0, m=0, s=0: excel_time(h, m, s),
    "HOUR": lambda c, x: int((to_number(x) % 1) * 24), "MINUTE": lambda c, x: int((to_number(x) % 1) * 1440) % 60,
    "NOW": lambda c: (datetime.utcnow().date() - EXCEL_EPOCH).days + (datetime.utcnow().hour * 3600 + datetime.utcnow().minute * 60 + datetime.utcnow().second) / 86400,
    "TODAY": lambda c: (date.today() - EXCEL_EPOCH).days, "EOMONTH": lambda c, s, m: eomonth(s, m), "EDATE": lambda c, s, m: edate(s, m),
    "NETWORKDAYS": lambda c, s, e, h=None: networkdays(s, e, h), "YEARFRAC": lambda c, s, e, basis=0: clean_number((serial_to_date(e) - serial_to_date(s)).days / 365), "WORKDAY": lambda c, s, d, h=None: workday(s, d, h),
    "HEX2DEC": lambda c, x: int(str(x), 16), "DEC2HEX": lambda c, x: format(int(to_number(x)), "X"), "BIN2DEC": lambda c, x: int(str(x), 2), "DEC2BIN": lambda c, x: format(int(to_number(x)), "b"),
    "OCT2DEC": lambda c, x: int(str(x), 8), "DEC2OCT": lambda c, x: format(int(to_number(x)), "o"),
    "BITAND": lambda c, a, b: int(to_number(a)) & int(to_number(b)), "BITOR": lambda c, a, b: int(to_number(a)) | int(to_number(b)), "BITXOR": lambda c, a, b: int(to_number(a)) ^ int(to_number(b)),
    "BITLSHIFT": lambda c, a, b: int(to_number(a)) << int(to_number(b)), "BITRSHIFT": lambda c, a, b: int(to_number(a)) >> int(to_number(b)),
    "DSUM": lambda c, db, f, cr: clean_number(sum(numbers(db_values(db, f, cr)))), "DAVERAGE": lambda c, db, f, cr: clean_number(sum(numbers(db_values(db, f, cr))) / len(numbers(db_values(db, f, cr)))),
    "DCOUNT": lambda c, db, f, cr: len(numbers(db_values(db, f, cr))), "DGET": lambda c, db, f, cr: (db_values(db, f, cr)[0] if len(db_values(db, f, cr)) == 1 else ERR_NUM),
    "DMAX": lambda c, db, f, cr: clean_number(max(numbers(db_values(db, f, cr)))), "DMIN": lambda c, db, f, cr: clean_number(min(numbers(db_values(db, f, cr)))),
    "VLOOKUP": func_vlookup, "HLOOKUP": func_hlookup, "MATCH": func_match, "INDEX": func_index, "XLOOKUP": func_xlookup,
    "OFFSET": lambda c, ref, rows, cols, height=1, width=1: RangeValue(c, ref.sheet if isinstance(ref, RangeValue) else c.sheet, (ref.r1 if isinstance(ref, RangeValue) else 1) + int(to_number(rows)), (ref.c1 if isinstance(ref, RangeValue) else 1) + int(to_number(cols)), (ref.r1 if isinstance(ref, RangeValue) else 1) + int(to_number(rows)) + int(to_number(height)) - 1, (ref.c1 if isinstance(ref, RangeValue) else 1) + int(to_number(cols)) + int(to_number(width)) - 1),
    "ADDRESS": lambda c, r, col, abs_num=1, a1=True, sheet=None: (str(sheet) + "!" if sheet else "") + f"${num_to_col(int(to_number(col)))}${int(to_number(r))}",
    "INDIRECT": lambda c, text: c.get_cell_value(*parse_sheet_ref(str(text).replace("$", ""), c.sheet)),
    "LEN": lambda c, x: len(str(x)), "LEFT": lambda c, x, n=1: str(x)[:int(to_number(n))], "RIGHT": lambda c, x, n=1: str(x)[-int(to_number(n)):],
    "MID": lambda c, x, start, n: str(x)[int(to_number(start)) - 1:int(to_number(start)) - 1 + int(to_number(n))],
    "UPPER": lambda c, x: str(x).upper(), "LOWER": lambda c, x: str(x).lower(), "PROPER": lambda c, x: str(x).title(), "TRIM": lambda c, x: " ".join(str(x).split()),
    "CONCAT": lambda c, *a: "".join(format_plain(x) for x in flatten(list(a))), "CONCATENATE": lambda c, *a: "".join(format_plain(x) for x in flatten(list(a))),
    "TEXTJOIN": lambda c, delim, ignore_empty, *a: textjoin(delim, ignore_empty, *a),
    "SUBSTITUTE": lambda c, text, old, new, inst=None: str(text).replace(str(old), str(new), 1 if inst is not None else -1),
    "REPT": lambda c, text, n: str(text) * int(to_number(n)),
    "FIND": lambda c, find, within, start=1: str(within).index(str(find), int(to_number(start)) - 1) + 1,
    "SEARCH": lambda c, find, within, start=1: str(within).lower().index(str(find).lower(), int(to_number(start)) - 1) + 1,
    "VALUE": lambda c, x: to_number(str(x).replace(",", "")),
    "TEXT": lambda c, x, fmt: excel_text(x, fmt),
    "CHOOSE": lambda c, index, *vals: choose(index, *vals),
    "TEXTSPLIT": textsplit, "TEXTBEFORE": textbefore, "TEXTAFTER": textafter, "UNICODE": lambda c, x: ord(str(x)[0]), "UNICHAR": lambda c, x: chr(int(to_number(x))),
    "AND": lambda c, *a: all(to_bool(x) for x in flatten(list(a))), "OR": lambda c, *a: any(to_bool(x) for x in flatten(list(a))), "NOT": lambda c, x: not to_bool(x),
    "ISNUMBER": lambda c, x: isinstance(scalar(x), (int, float)) and not isinstance(scalar(x), bool), "ISTEXT": lambda c, x: isinstance(scalar(x), str) and not is_error(scalar(x)),
    "ISERROR": lambda c, x: is_error(scalar(x)), "ISBLANK": lambda c, x: scalar(x) in (None, ""),
    "SEQUENCE": func_sequence, "MAP": func_map, "REDUCE": func_reduce, "FILTER": func_filter, "SORT": func_sort, "UNIQUE": func_unique, "BYROW": func_byrow, "BYCOL": func_bycol,
}


def make_cell(ref: str, cell: dict[str, Any] | None, ghost: dict[str, Any] | None = None, wb: dict[str, Any] | None = None, sheet: str | None = None) -> dict[str, Any]:
    if ghost:
        out = {"sheet": sheet, "ref": ref, "input": None, "value": ghost.get("value"), "kind": "spill", "anchor": ghost.get("anchor"), "display": format_display(ghost.get("value"), None)}
        return {k: v for k, v in out.items() if v is not None}
    if not cell:
        return {"sheet": sheet, "ref": ref, "input": None, "value": None, "kind": "empty", "display": ""}
    value = cell.get("value")
    out = {
        "sheet": sheet,
        "ref": ref,
        "input": cell.get("input"),
        "value": value,
        "kind": cell.get("kind") or cell_kind(value),
        "display": format_display(value, cell.get("format")),
    }
    if cell.get("spill"):
        out["spill"] = True
        out["spill_range"] = cell.get("spill_range")
    if cell.get("format"):
        out["format"] = cell.get("format")
    if cell.get("style"):
        out["style"] = cell.get("style")
    if wb and sheet:
        cf = cell_conditional_formats(wb, sheet, ref, value)
        if cf:
            out["cf"] = cf
    return out


def format_display(value: Any, fmt: str | None) -> str:
    if value is None:
        return ""
    if is_error(value):
        return value
    if fmt and isinstance(value, (int, float)) and not isinstance(value, bool):
        if "%" in fmt:
            decimals = 0
            m = re.search(r"0\.(0+)%", fmt)
            if m:
                decimals = len(m.group(1))
            return f"{value * 100:.{decimals}f}%"
        if "$" in fmt:
            decimals = 2 if ".00" in fmt else 0
            return "$" + f"{value:,.{decimals}f}"
        if "yyyy" in fmt.lower():
            d = serial_to_date(value)
            return d.strftime("%Y-%m-%d")
        m = re.search(r"0\.(0+)", fmt)
        if m:
            return f"{value:.{len(m.group(1))}f}"
    return format_plain(value)


def build_ghosts(wb: dict[str, Any]) -> dict[str, dict[str, dict[str, Any]]]:
    ghosts: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for sname, sheet in wb["sheets"].items():
        for ref, cell in sheet["cells"].items():
            if cell.get("spill") and cell.get("spill_values"):
                vals = cell["spill_values"]
                r0, c0 = split_ref(ref)
                for i, row in enumerate(vals):
                    for j, value in enumerate(row):
                        target = ref_from_rc(r0 + i, c0 + j)
                        if target == ref:
                            continue
                        ghosts[sname][target] = {"anchor": ref, "value": value}
    return ghosts


def all_formula_cells(wb: dict[str, Any]) -> list[tuple[str, str]]:
    out = []
    for sname, sheet in wb["sheets"].items():
        for ref, cell in sheet["cells"].items():
            if isinstance(cell.get("input"), str) and cell["input"].startswith("="):
                out.append((sname, ref))
    return out


def recompute(wb: dict[str, Any], changed: set[tuple[str, str]] | None = None) -> list[dict[str, Any]]:
    # Clear prior spill materialization metadata.
    for sheet in wb["sheets"].values():
        for cell in sheet["cells"].values():
            cell.pop("spill", None)
            cell.pop("spill_range", None)
            cell.pop("spill_values", None)
            if isinstance(cell.get("input"), str) and cell["input"].startswith("="):
                cell["deps"] = [list(x) for x in sorted(extract_deps(cell["input"], sheet["name"], wb))]

    formulas = all_formula_cells(wb)
    deps = {key: {tuple(x) for x in wb["sheets"][key[0]]["cells"][key[1]].get("deps", [])} for key in formulas}
    dependents: dict[tuple[str, str], set[tuple[str, str]]] = defaultdict(set)
    for cell_key, ds in deps.items():
        for d in ds:
            dependents[d].add(cell_key)

    if changed is None:
        dirty = set(formulas)
    else:
        dirty = set(changed)
        q = deque(changed)
        while q:
            k = q.popleft()
            for dep in dependents.get(k, set()):
                if dep not in dirty:
                    dirty.add(dep)
                    q.append(dep)
        dirty |= {k for k in changed if k in [(s, r) for s, sh in wb["sheets"].items() for r in sh["cells"]]}

    values: dict[tuple[str, str], Any] = {}
    for sname, sheet in wb["sheets"].items():
        for ref, cell in sheet["cells"].items():
            if not (isinstance(cell.get("input"), str) and cell["input"].startswith("=")):
                values[(sname, ref)] = cell.get("value")
            else:
                values[(sname, ref)] = cell.get("value")

    iterative = wb.get("settings", {}).get("iterative_calc", {})
    if iterative.get("enabled"):
        formula_set = set(formulas)
        max_iter = int(iterative.get("max_iterations", 100))
        max_change = float(iterative.get("max_change", 0.001))
        for _ in range(max_iter):
            biggest = 0.0
            for key in formulas:
                sname, ref = key
                cell = wb["sheets"][sname]["cells"][ref]
                old = values.get(key)
                val = EvalContext(wb, sname, values).eval_formula(cell["input"])
                values[key] = scalar(val)
                if isinstance(old, (int, float)) and isinstance(values[key], (int, float)):
                    biggest = max(biggest, abs(float(old) - float(values[key])))
                elif old != values[key]:
                    biggest = max(biggest, 1.0)
            if biggest <= max_change:
                break
    else:
        temp, perm, cycle_nodes = set(), set(), set()
        order: list[tuple[str, str]] = []

        def visit(k):
            if k in perm:
                return
            if k in temp:
                cycle_nodes.update(temp)
                return
            temp.add(k)
            for d in deps.get(k, set()):
                if d in deps:
                    visit(d)
            temp.remove(k)
            perm.add(k)
            order.append(k)

        for k in formulas:
            visit(k)
        for key in order:
            if changed is not None and key not in dirty:
                continue
            sname, ref = key
            cell = wb["sheets"][sname]["cells"][ref]
            if key in cycle_nodes:
                values[key] = ERR_CIRC
            else:
                values[key] = EvalContext(wb, sname, values).eval_formula(cell["input"])

    changed_cells: list[dict[str, Any]] = []
    # Materialize formulas and spills. Anchors check occupied target cells.
    for sname, ref in formulas:
        cell = wb["sheets"][sname]["cells"][ref]
        raw = values.get((sname, ref))
        arr = array2d(raw)
        is_array = isinstance(raw, (list, RangeValue)) and (len(arr) > 1 or (arr and len(arr[0]) > 1))
        if is_array:
            r0, c0 = split_ref(ref)
            blocked = False
            for i, row in enumerate(arr):
                for j, _ in enumerate(row):
                    target = ref_from_rc(r0 + i, c0 + j)
                    if target == ref:
                        continue
                    other = wb["sheets"][sname]["cells"].get(target)
                    if other and other.get("input") not in (None, ""):
                        blocked = True
            if blocked:
                cell["value"], cell["kind"] = ERR_SPILL, "error"
            else:
                cell["value"] = arr[0][0] if arr and arr[0] else None
                cell["kind"] = cell_kind(cell["value"])
                cell["spill"] = True
                cell["spill_range"] = {"rows": len(arr), "cols": max(len(r) for r in arr) if arr else 0}
                cell["spill_values"] = arr
        else:
            cell["value"] = scalar(raw)
            cell["kind"] = cell_kind(cell["value"])

    ghosts = build_ghosts(wb)
    include: set[tuple[str, str]] = set()
    if changed is None:
        include = set((s, r) for s, sh in wb["sheets"].items() for r in sh["cells"])
    else:
        include = dirty | changed
    for sname, ref in sorted(include):
        if sname in wb["sheets"]:
            cell = wb["sheets"][sname]["cells"].get(ref)
            changed_cells.append(make_cell(ref, cell, None, wb, sname))
    for sname, gs in ghosts.items():
        for ref, ghost in gs.items():
            changed_cells.append(make_cell(ref, None, ghost, wb, sname))
    wb["updated"] = now_ms()
    return changed_cells


def get_visible_cell(wb: dict[str, Any], sheet: str, ref: str) -> dict[str, Any]:
    if sheet not in wb["sheets"]:
        raise ApiError(404, "NotFound")
    ref = norm_ref(ref)
    ghosts = build_ghosts(wb)
    return make_cell(ref, wb["sheets"][sheet]["cells"].get(ref), ghosts.get(sheet, {}).get(ref), wb, sheet)


def workbook_summary(wb: dict[str, Any], full: bool = True) -> dict[str, Any]:
    out = {"id": wb["id"], "name": wb["name"]}
    if full:
        ghosts = build_ghosts(wb)
        sheets = []
        for sname in wb.get("sheet_order", list(wb["sheets"])):
            sheet = wb["sheets"][sname]
            refs = set(sheet["cells"]) | set(ghosts.get(sname, {}))
            sheets.append({"name": sname, "cells": [make_cell(ref, sheet["cells"].get(ref), ghosts.get(sname, {}).get(ref), wb, sname) for ref in sorted(refs, key=lambda x: split_ref(x))]})
        out["sheets"] = sheets
    else:
        out["sheets"] = wb.get("sheet_order", list(wb["sheets"]))
    return out


def validate_literal(wb: dict[str, Any], sheet: str, ref: str, input_value: Any):
    if input_value is None or (isinstance(input_value, str) and input_value.startswith("=")):
        return
    val, _ = coerce_literal(input_value)
    for rule_entry in wb.get("validations", []):
        if rule_entry.get("sheet") != sheet:
            continue
        if ref not in range_refs(rule_entry.get("range", ref)):
            continue
        rule = rule_entry.get("rule", {})
        kind = rule.get("kind")
        ok = True
        if kind == "list":
            ok = str(val) in {str(x) for x in rule.get("values", [])}
        elif kind in ("integer", "decimal"):
            try:
                n = to_number(val)
                if kind == "integer":
                    ok = abs(n - round(n)) < 1e-9
                op = rule.get("op")
                if op == "between":
                    ok = ok and float(rule.get("min", -math.inf)) <= n <= float(rule.get("max", math.inf))
                elif op:
                    ok = ok and criteria_match(n, op + str(rule.get("value", rule.get("min", 0))))
            except Exception:
                ok = False
        elif kind == "text_length":
            n = len(str(val))
            op = rule.get("op")
            if op == "between":
                ok = int(rule.get("min", 0)) <= n <= int(rule.get("max", 10**9))
            elif op:
                ok = criteria_match(n, op + str(rule.get("value", 0)))
        if not ok:
            raise ApiError(400, "ValidationError")


def apply_patches(wb: dict[str, Any], patches: list[dict[str, Any]], locale: str | None = None) -> list[dict[str, Any]]:
    ghosts = build_ghosts(wb)
    changed: set[tuple[str, str]] = set()
    locale = locale or wb.get("settings", {}).get("locale", "en-US")
    for p in patches:
        sheet = p.get("sheet")
        ref = norm_ref(p.get("ref"))
        if sheet not in wb["sheets"]:
            raise ApiError(404, "NotFound")
        if ref in ghosts.get(sheet, {}):
            raise ApiError(400, "SpillTargetWriteError")
        inp = p.get("input")
        validate_literal(wb, sheet, ref, inp)
        cells = wb["sheets"][sheet]["cells"]
        if inp is None:
            if ref in cells:
                old = cells[ref]
                fmt, sty = old.get("format"), old.get("style")
                if p.get("format") or p.get("style"):
                    cells[ref] = {"input": None, "value": None, "kind": "empty"}
                else:
                    cells.pop(ref, None)
                if fmt and ref in cells:
                    cells[ref]["format"] = fmt
                if sty and ref in cells:
                    cells[ref]["style"] = sty
            changed.add((sheet, ref))
            continue
        inp = translate_formula(inp, locale)
        if isinstance(inp, str) and inp.startswith("="):
            try:
                parse_formula(inp)
            except ApiError:
                raise
            cell = cells.setdefault(ref, {})
            cell.update({"input": inp, "kind": "empty", "value": None})
        else:
            val, kind = coerce_literal(inp)
            cell = cells.setdefault(ref, {})
            cell.update({"input": inp, "value": val, "kind": kind})
        if "format" in p:
            cell["format"] = p["format"]
        if "style" in p:
            cell["style"] = p["style"]
        changed.add((sheet, ref))
    return recompute(wb, changed)


def cell_conditional_formats(wb: dict[str, Any], sheet: str, ref: str, value: Any) -> list[dict[str, Any]]:
    out = []
    for cf in wb.get("conditional_formats", []):
        if cf.get("sheet") != sheet or ref not in range_refs(cf.get("range", ref)):
            continue
        rule = cf.get("rule", {})
        if rule.get("kind") != "cell_value":
            continue
        op = rule.get("op", "=")
        if op == "between":
            lo, hi = rule.get("value", [None, None])
            ok = to_number(lo) <= to_number(value) <= to_number(hi)
        else:
            ok = criteria_match(value, op + str(rule.get("value", "")))
        if ok:
            out.append({"style": rule.get("style", {})})
    return out


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.get("/")
async def root():
    return HTMLResponse(SPA_HTML)


@app.get("/api/workbooks")
async def list_workbooks():
    return {"workbooks": [workbook_summary(wb, full=False) for wb in store.state["workbooks"].values()]}


@app.post("/api/workbooks")
async def create_workbook(body: dict[str, Any]):
    async with store.lock:
        wb_id = int(store.state.get("next_id", 1))
        store.state["next_id"] = wb_id + 1
        wb = default_workbook(wb_id, body.get("name") or f"Workbook {wb_id}")
        store.state["workbooks"][str(wb_id)] = wb
        store.save()
        return workbook_summary(wb, full=False)


@app.get("/api/workbooks/{wb_id}")
async def get_workbook(wb_id: int):
    return workbook_summary(store.get(wb_id), full=True)


@app.delete("/api/workbooks/{wb_id}")
async def delete_workbook(wb_id: int):
    async with store.lock:
        if str(wb_id) not in store.state["workbooks"]:
            raise ApiError(404, "NotFound")
        del store.state["workbooks"][str(wb_id)]
        store.save()
    return {"ok": True}


@app.post("/api/workbooks/{wb_id}/sheets")
async def add_sheet(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        name = body.get("name")
        if not name:
            raise ApiError(400, "ValidationError")
        if name not in wb["sheets"]:
            wb["sheets"][name] = {"name": name, "cells": {}}
            wb.setdefault("sheet_order", []).append(name)
        store.save()
        return {"name": name}


@app.delete("/api/workbooks/{wb_id}/sheets/{sheet}")
async def remove_sheet(wb_id: int, sheet: str):
    async with store.lock:
        wb = store.get(wb_id)
        if sheet not in wb["sheets"]:
            raise ApiError(404, "NotFound")
        if len(wb["sheets"]) == 1:
            raise ApiError(400, "ValidationError")
        del wb["sheets"][sheet]
        wb["sheet_order"] = [s for s in wb.get("sheet_order", []) if s != sheet]
        recompute(wb, None)
        store.save()
        return {"ok": True}


@app.post("/api/workbooks/{wb_id}/cells")
async def patch_cells(wb_id: int, request: Request, locale: str | None = Query(default=None)):
    body = await request.json()
    async with store.lock:
        wb = store.get(wb_id)
        cells = apply_patches(wb, body.get("patches", []), locale)
        store.save()
        return {"cells": cells}


@app.get("/api/workbooks/{wb_id}/cells/{sheet}/{ref}")
async def read_cell(wb_id: int, sheet: str, ref: str):
    return get_visible_cell(store.get(wb_id), sheet, ref)


def rewrite_formula_refs(formula: str, dr: int, dc: int) -> str:
    def repl(m: re.Match) -> str:
        token = m.group(1)
        mm = re.match(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)$", token)
        if not mm:
            return token
        col_abs, col, row_abs, row = mm.groups()
        r = int(row)
        c = col_to_num(col)
        if not row_abs:
            r += dr
        if not col_abs:
            c += dc
        return f"{col_abs}{num_to_col(max(1, c))}{row_abs}{max(1, r)}"

    def skip_strings(text: str) -> str:
        parts = re.split(r'("(?:[^"]|"")*")', text)
        for i in range(0, len(parts), 2):
            parts[i] = CELL_FIND_RE.sub(repl, parts[i])
        return "".join(parts)

    return "=" + skip_strings(formula[1:]) if formula.startswith("=") else formula


@app.post("/api/workbooks/{wb_id}/fill")
async def fill_cells(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        sheet = body.get("sheet")
        if sheet not in wb["sheets"]:
            raise ApiError(404, "NotFound")
        source = norm_ref(body.get("source"))
        src_cell = wb["sheets"][sheet]["cells"].get(source)
        if not src_cell:
            return {"cells": []}
        sr, sc = split_ref(source)
        patches = []
        for ref in range_refs(body.get("targets")):
            tr, tc = split_ref(ref)
            inp = src_cell.get("input")
            if isinstance(inp, str) and inp.startswith("="):
                inp = rewrite_formula_refs(inp, tr - sr, tc - sc)
            p = {"sheet": sheet, "ref": ref, "input": inp}
            if src_cell.get("format"):
                p["format"] = src_cell.get("format")
            if src_cell.get("style"):
                p["style"] = src_cell.get("style")
            patches.append(p)
        cells = apply_patches(wb, patches)
        store.save()
        return {"cells": cells}


@app.post("/api/workbooks/{wb_id}/sort")
async def sort_range(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        sheet = body.get("sheet")
        if sheet not in wb["sheets"]:
            raise ApiError(404, "NotFound")
        r1, c1, r2, c2 = rect_bounds(body.get("range"))
        rows = []
        for r in range(r1, r2 + 1):
            row = []
            for c in range(c1, c2 + 1):
                ref = ref_from_rc(r, c)
                cell = deepcopy(wb["sheets"][sheet]["cells"].get(ref, {"input": None, "value": None, "kind": "empty"}))
                row.append(cell)
            rows.append(row)
        by = body.get("by", [])
        for spec in reversed(by):
            idx = int(spec.get("column", 0))
            asc = bool(spec.get("asc", True))
            rows.sort(key=lambda row: (row[idx].get("value") is None, row[idx].get("value")), reverse=not asc)
        patches = []
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                patches.append({"sheet": sheet, "ref": ref_from_rc(r1 + i, c1 + j), "input": cell.get("input"), "format": cell.get("format"), "style": cell.get("style")})
        cells = apply_patches(wb, patches)
        store.save()
        return {"cells": cells}


@app.post("/api/workbooks/{wb_id}/filter")
async def filter_range(wb_id: int, body: dict[str, Any]):
    wb = store.get(wb_id)
    sheet = body.get("sheet")
    if sheet not in wb["sheets"]:
        raise ApiError(404, "NotFound")
    r1, c1, r2, c2 = rect_bounds(body.get("range"))
    col_idx = int(body.get("criteria", {}).get("column", 0))
    expr = body.get("criteria", {}).get("expr", "")
    rows = []
    for r in range(r1, r2 + 1):
        cells = [get_visible_cell(wb, sheet, ref_from_rc(r, c)) for c in range(c1, c2 + 1)]
        if col_idx < len(cells) and criteria_match(cells[col_idx].get("value"), expr):
            rows.append({"row": r, "cells": cells})
    return {"rows": rows}


@app.get("/api/workbooks/{wb_id}/csv")
async def export_csv(wb_id: int, sheet: str):
    wb = store.get(wb_id)
    if sheet not in wb["sheets"]:
        raise ApiError(404, "NotFound")
    refs = set(wb["sheets"][sheet]["cells"]) | set(build_ghosts(wb).get(sheet, {}))
    max_r = max((split_ref(r)[0] for r in refs), default=1)
    max_c = max((split_ref(r)[1] for r in refs), default=1)
    out = io.StringIO()
    writer = csv.writer(out)
    for r in range(1, max_r + 1):
        writer.writerow([get_visible_cell(wb, sheet, ref_from_rc(r, c)).get("value") for c in range(1, max_c + 1)])
    return Response(out.getvalue(), media_type="text/csv")


@app.post("/api/workbooks/{wb_id}/csv")
async def import_csv(wb_id: int, sheet: str, file: UploadFile = File(...)):
    text = (await file.read()).decode("utf-8")
    async with store.lock:
        wb = store.get(wb_id)
        if sheet not in wb["sheets"]:
            wb["sheets"][sheet] = {"name": sheet, "cells": {}}
            wb.setdefault("sheet_order", []).append(sheet)
        wb["sheets"][sheet]["cells"] = {}
        patches = []
        for i, row in enumerate(csv.reader(io.StringIO(text)), 1):
            for j, val in enumerate(row, 1):
                patches.append({"sheet": sheet, "ref": ref_from_rc(i, j), "input": val})
        cells = apply_patches(wb, patches)
        store.save()
        return {"cells": cells}


def import_xlsx_into(wb: dict[str, Any], data: bytes):
    xlb = load_workbook(io.BytesIO(data), data_only=False)
    wb["sheets"] = {}
    wb["sheet_order"] = list(xlb.sheetnames)
    for ws in xlb.worksheets:
        sheet = {"name": ws.title, "cells": {}}
        for row in ws.iter_rows():
            for xlcell in row:
                val = xlcell.value
                if val is None:
                    continue
                ref = xlcell.coordinate
                inp = val
                if isinstance(val, str) and val.startswith("="):
                    inp = val
                cell_val, kind = coerce_literal(inp)
                cell = {"input": inp, "value": cell_val, "kind": kind}
                if xlcell.number_format and xlcell.number_format != "General":
                    cell["format"] = xlcell.number_format
                style = {}
                if xlcell.font:
                    if xlcell.font.bold:
                        style["bold"] = True
                    if xlcell.font.italic:
                        style["italic"] = True
                    if xlcell.font.color and xlcell.font.color.type == "rgb" and xlcell.font.color.rgb:
                        style["font_color"] = "#" + xlcell.font.color.rgb[-6:]
                if xlcell.fill and xlcell.fill.fill_type == "solid" and xlcell.fill.fgColor and xlcell.fill.fgColor.rgb:
                    style["bg_color"] = "#" + xlcell.fill.fgColor.rgb[-6:]
                if style:
                    cell["style"] = style
                sheet["cells"][ref] = cell
        wb["sheets"][ws.title] = sheet
    wb["names"] = []
    try:
        for name, defn in xlb.defined_names.items():
            dests = list(defn.destinations)
            if dests:
                sh, rng = dests[0]
                wb["names"].append({"scope": "workbook", "name": name, "range": f"{sh}!{rng}"})
    except Exception:
        pass
    recompute(wb, None)


@app.post("/api/workbooks/{wb_id}/xlsx")
async def import_xlsx(wb_id: int, file: UploadFile = File(...)):
    data = await file.read()
    async with store.lock:
        wb = store.get(wb_id)
        import_xlsx_into(wb, data)
        store.save()
        return workbook_summary(wb, full=True)


def export_xlsx_bytes(wb: dict[str, Any]) -> bytes:
    xlb = XLWorkbook()
    default = xlb.active
    xlb.remove(default)
    for sname in wb.get("sheet_order", list(wb["sheets"])):
        ws = xlb.create_sheet(sname)
        sheet = wb["sheets"][sname]
        ghosts = build_ghosts(wb).get(sname, {})
        refs = set(sheet["cells"]) | set(ghosts)
        for ref in refs:
            cell = sheet["cells"].get(ref)
            ghost = ghosts.get(ref)
            ws_cell = ws[ref]
            if cell:
                inp = cell.get("input")
                ws_cell.value = inp if isinstance(inp, str) and inp.startswith("=") else cell.get("value")
                if cell.get("format"):
                    ws_cell.number_format = cell["format"]
                st = cell.get("style") or {}
                if st:
                    color = (st.get("font_color") or "#000000").replace("#", "")
                    ws_cell.font = Font(bold=bool(st.get("bold")), italic=bool(st.get("italic")), color=color)
                    if st.get("bg_color"):
                        ws_cell.fill = PatternFill("solid", fgColor=st["bg_color"].replace("#", ""))
            elif ghost:
                ws_cell.value = ghost.get("value")
        for cf in wb.get("conditional_formats", []):
            if cf.get("sheet") == sname:
                rule = cf.get("rule", {})
                style = rule.get("style", {})
                fill = PatternFill("solid", fgColor=(style.get("bg_color", "#FFFF00").replace("#", ""))) if style.get("bg_color") else None
                op_map = {"=": "equal", "<>": "notEqual", "<": "lessThan", ">": "greaterThan", "<=": "lessThanOrEqual", ">=": "greaterThanOrEqual", "between": "between"}
                vals = rule.get("value")
                formula = vals if isinstance(vals, list) else [str(vals)]
                ws.conditional_formatting.add(cf["range"], CellIsRule(operator=op_map.get(rule.get("op"), "equal"), formula=formula, fill=fill))
    for n in wb.get("names", []):
        xlb.defined_names.add(DefinedName(n["name"], attr_text=n["range"]))
    bio = io.BytesIO()
    xlb.save(bio)
    return inject_xlsx_cached_values(bio.getvalue(), wb)


def inject_xlsx_cached_values(blob: bytes, wb: dict[str, Any]) -> bytes:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", ns["main"])
    src = io.BytesIO(blob)
    dst = io.BytesIO()
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        sheet_order = wb.get("sheet_order", list(wb["sheets"]))
        sheet_xml = {f"xl/worksheets/sheet{i + 1}.xml": s for i, s in enumerate(sheet_order)}
        for item in zin.infolist():
            data = zin.read(item.filename)
            sname = sheet_xml.get(item.filename)
            if sname:
                try:
                    root = ET.fromstring(data)
                    values = {}
                    for ref, cell in wb["sheets"][sname]["cells"].items():
                        if isinstance(cell.get("input"), str) and cell["input"].startswith("="):
                            values[ref] = cell.get("value")
                    for cnode in root.findall(".//main:c", ns):
                        ref = cnode.attrib.get("r")
                        if ref not in values:
                            continue
                        if cnode.find("main:f", ns) is None:
                            continue
                        for vnode in list(cnode.findall("main:v", ns)):
                            cnode.remove(vnode)
                        val = values[ref]
                        vnode = ET.Element(f"{{{ns['main']}}}v")
                        if isinstance(val, bool):
                            cnode.attrib["t"] = "b"
                            vnode.text = "1" if val else "0"
                        elif isinstance(val, (int, float)) and not isinstance(val, bool):
                            cnode.attrib.pop("t", None)
                            vnode.text = str(val)
                        elif val is None:
                            vnode.text = ""
                        elif is_error(val):
                            cnode.attrib["t"] = "e"
                            vnode.text = val
                        else:
                            cnode.attrib["t"] = "str"
                            vnode.text = str(val)
                        cnode.append(vnode)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                except Exception:
                    pass
            zout.writestr(item, data)
    return dst.getvalue()


@app.get("/api/workbooks/{wb_id}/xlsx")
async def export_xlsx(wb_id: int):
    data = export_xlsx_bytes(store.get(wb_id))
    return Response(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/workbooks/{wb_id}/settings")
async def get_settings(wb_id: int):
    return store.get(wb_id).get("settings", {})


@app.put("/api/workbooks/{wb_id}/settings")
async def put_settings(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        wb.setdefault("settings", {}).update(body)
        if "iterative_calc" in body:
            wb["settings"]["iterative_calc"] = {**{"enabled": False, "max_iterations": 100, "max_change": 0.001}, **body["iterative_calc"]}
        recompute(wb, None)
        store.save()
        return wb["settings"]


@app.post("/api/workbooks/{wb_id}/names")
async def add_name(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        entry = {"scope": body.get("scope", "workbook"), "sheet": body.get("sheet"), "name": body.get("name"), "range": body.get("range")}
        wb.setdefault("names", []).append(entry)
        recompute(wb, None)
        store.save()
        return entry


@app.get("/api/workbooks/{wb_id}/names")
async def get_names(wb_id: int):
    return {"names": store.get(wb_id).get("names", [])}


@app.post("/api/workbooks/{wb_id}/conditional_formats")
async def add_cf(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        wb.setdefault("conditional_formats", []).append(body)
        store.save()
        return body


@app.post("/api/workbooks/{wb_id}/data_validation")
async def add_validation(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        wb.setdefault("validations", []).append(body)
        store.save()
        return body


@app.post("/api/workbooks/{wb_id}/pivot")
async def pivot(wb_id: int, body: dict[str, Any]):
    wb = store.get(wb_id)
    sheet, rng = parse_sheet_ref(body.get("source"), "Sheet1")
    if sheet not in wb["sheets"]:
        raise ApiError(404, "NotFound")
    r1, c1, r2, c2 = rect_bounds(rng)
    headers = [get_visible_cell(wb, sheet, ref_from_rc(r1, c)).get("value") for c in range(c1, c2 + 1)]
    data = []
    for r in range(r1 + 1, r2 + 1):
        row = {headers[c - c1]: get_visible_cell(wb, sheet, ref_from_rc(r, c)).get("value") for c in range(c1, c2 + 1)}
        data.append(row)
    row_fields = body.get("rows", [])
    col_fields = body.get("cols", [])
    val_items = list((body.get("values") or {}).items())
    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in data:
        key = tuple(row.get(f) for f in row_fields + col_fields)
        groups[key].append(row)
    cells = []
    for key in sorted(groups, key=lambda x: tuple(str(v) for v in x)):
        rows = groups[key]
        for col, agg in val_items:
            vals = [r.get(col) for r in rows]
            nums = [to_number(v) for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if agg == "sum":
                val = sum(nums)
            elif agg == "avg":
                val = sum(nums) / len(nums) if nums else None
            elif agg == "count":
                val = len([v for v in vals if v not in (None, "")])
            elif agg == "min":
                val = min(nums) if nums else None
            elif agg == "max":
                val = max(nums) if nums else None
            else:
                val = None
            cells.append({"row": list(key[:len(row_fields)]), "col": list(key[len(row_fields):]), "field": col, "agg": agg, "value": clean_number(val) if isinstance(val, float) else val})
    return {"cells": cells}


@app.post("/api/workbooks/{wb_id}/goal_seek")
async def goal_seek(wb_id: int, body: dict[str, Any]):
    async with store.lock:
        wb = store.get(wb_id)
        target_sheet, target_ref = parse_sheet_ref(body["target_cell"], "Sheet1")
        change_sheet, change_ref = parse_sheet_ref(body["changing_cell"], "Sheet1")
        target_ref, change_ref = norm_ref(target_ref), norm_ref(change_ref)
        target_value = float(body["target_value"])
        tol = float(body.get("tol", 1e-4))
        max_iter = int(body.get("max_iter", 80))
        original = deepcopy(wb["sheets"][change_sheet]["cells"].get(change_ref))

        def eval_at(x: float) -> float:
            apply_patches(wb, [{"sheet": change_sheet, "ref": change_ref, "input": str(x)}])
            v = get_visible_cell(wb, target_sheet, target_ref).get("value")
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                raise ValueError("target not numeric")
            return float(v) - target_value

        try:
            guesses = [0.0, 1.0, -1.0, 10.0, -10.0, 100.0, -100.0, 1000.0, -1000.0]
            root = None
            last_x, last_y = guesses[0], eval_at(guesses[0])
            for x in guesses[1:]:
                y = eval_at(x)
                if abs(y) <= tol:
                    root = x
                    break
                if last_y == 0 or y == 0 or (last_y < 0 < y) or (y < 0 < last_y):
                    root = optimize.brentq(lambda z: eval_at(z), last_x, x, xtol=tol, maxiter=max_iter)
                    break
                last_x, last_y = x, y
            if root is None:
                root = optimize.newton(lambda z: eval_at(z), 1.0, maxiter=max_iter, tol=tol)
            output_delta = eval_at(root)
            if abs(output_delta) > max(tol, 1e-3):
                raise ValueError("not converged")
            cells = apply_patches(wb, [{"sheet": change_sheet, "ref": change_ref, "input": str(root)}])
            store.save()
            return {"converged": True, "iterations": max_iter, "input": root, "output": target_value + output_delta, "cells": cells}
        except Exception:
            if original is None:
                wb["sheets"][change_sheet]["cells"].pop(change_ref, None)
            else:
                wb["sheets"][change_sheet]["cells"][change_ref] = original
            recompute(wb, None)
            store.save()
            raise ApiError(400, "NotConverged")


collab_clients: dict[int, set[WebSocket]] = defaultdict(set)


async def broadcast(wb_id: int, msg: dict[str, Any]):
    dead = []
    for ws in list(collab_clients.get(wb_id, set())):
        try:
            await ws.send_json(msg)
        except Exception:
            dead.append(ws)
    for ws in dead:
        collab_clients[wb_id].discard(ws)


@app.websocket("/api/workbooks/{wb_id}/collab")
async def collab(ws: WebSocket, wb_id: int):
    await ws.accept()
    actor = "anonymous"
    collab_clients[wb_id].add(ws)
    try:
        first = await ws.receive_json()
        if first.get("type") != "hello":
            await ws.close()
            return
        actor = first.get("actor") or actor
        wb = store.get(wb_id)
        await ws.send_json({"type": "welcome", "seq": wb.get("seq", 0), "snapshot": workbook_summary(wb, full=True)})
        since = first.get("since_seq")
        if since is not None:
            for ev in wb.get("events", []):
                if ev.get("seq", 0) > int(since):
                    await ws.send_json({"type": "event", **ev})
        while True:
            msg = await ws.receive_json()
            typ = msg.get("type")
            if typ == "presence":
                await broadcast(wb_id, {"type": "presence_event", "actor": actor, "sheet": msg.get("sheet"), "ref": msg.get("ref"), "online": True})
            elif typ == "op":
                async with store.lock:
                    wb = store.get(wb_id)
                    op = msg.get("op", {})
                    kind = op.get("kind")
                    cells: list[dict[str, Any]] = []
                    if kind == "set":
                        cells = apply_patches(wb, [{"sheet": op.get("sheet"), "ref": op.get("ref"), "input": op.get("input")}])
                    elif kind == "clear":
                        cells = apply_patches(wb, [{"sheet": op.get("sheet"), "ref": op.get("ref"), "input": None}])
                    elif kind == "add_sheet":
                        name = op.get("name")
                        if name not in wb["sheets"]:
                            wb["sheets"][name] = {"name": name, "cells": {}}
                            wb.setdefault("sheet_order", []).append(name)
                    elif kind == "remove_sheet":
                        name = op.get("name") or op.get("sheet")
                        if name in wb["sheets"] and len(wb["sheets"]) > 1:
                            del wb["sheets"][name]
                            wb["sheet_order"] = [s for s in wb["sheet_order"] if s != name]
                    wb["seq"] = int(wb.get("seq", 0)) + 1
                    ev = {"seq": wb["seq"], "actor": actor, "client_seq": msg.get("client_seq"), "op": op, "cells": cells}
                    wb.setdefault("events", []).append(ev)
                    wb["events"] = wb["events"][-1000:]
                    store.save()
                await ws.send_json({"type": "ack", "seq": ev["seq"], "client_seq": msg.get("client_seq")})
                await broadcast(wb_id, {"type": "event", **ev})
    except WebSocketDisconnect:
        pass
    finally:
        collab_clients[wb_id].discard(ws)
        try:
            await broadcast(wb_id, {"type": "presence_event", "actor": actor, "sheet": None, "ref": None, "online": False})
        except Exception:
            pass


SPA_HTML = r"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Tabula</title>
  <style>
    :root { color-scheme: light; font-family: Inter, ui-sans-serif, system-ui, -apple-system, Segoe UI, sans-serif; }
    body { margin: 0; height: 100vh; display: grid; grid-template-columns: 240px 1fr; background: #f5f7fb; color: #172033; }
    aside { background: #ffffff; border-right: 1px solid #d8dde8; padding: 14px; overflow: auto; }
    main { display: grid; grid-template-rows: 48px 1fr 42px; min-width: 0; }
    .topbar { display: flex; align-items: center; gap: 10px; padding: 8px 12px; background: #ffffff; border-bottom: 1px solid #d8dde8; }
    input[data-testid="formula-bar"] { flex: 1; height: 30px; border: 1px solid #b9c1d1; border-radius: 4px; padding: 0 10px; font: 14px ui-monospace, SFMono-Regular, Menlo, monospace; }
    button { border: 1px solid #b9c1d1; background: #ffffff; border-radius: 4px; height: 30px; padding: 0 10px; cursor: pointer; }
    button:hover { background: #eef3fa; }
    .brand { font-weight: 700; font-size: 18px; margin-bottom: 12px; }
    #gridWrap { overflow: auto; padding: 12px; }
    table { border-collapse: collapse; background: #ffffff; box-shadow: 0 0 0 1px #d8dde8; }
    th, td { border: 1px solid #d8dde8; min-width: 96px; height: 28px; padding: 0 6px; font-size: 13px; white-space: nowrap; }
    th { background: #eef2f7; color: #475569; text-align: center; font-weight: 600; position: sticky; top: 0; z-index: 1; }
    th.rowhead { min-width: 42px; left: 0; z-index: 2; }
    td { cursor: cell; }
    td.active { outline: 2px solid #2563eb; outline-offset: -2px; }
    td input { width: 100%; height: 100%; border: 0; outline: 0; padding: 0; font: inherit; }
    .tabs { display: flex; gap: 6px; align-items: center; padding: 6px 12px; background: #ffffff; border-top: 1px solid #d8dde8; }
    .tab.active, .wb.active { background: #dbeafe; border-color: #60a5fa; }
    .wb { width: 100%; text-align: left; margin: 3px 0; overflow: hidden; text-overflow: ellipsis; }
  </style>
</head>
<body>
  <aside>
    <div class="brand">Tabula</div>
    <button data-testid="new-workbook-btn" id="newBtn">New workbook</button>
    <div data-testid="workbook-list" id="workbookList" style="margin-top:12px"></div>
  </aside>
  <main>
    <div class="topbar">
      <strong id="activeRef">A1</strong>
      <input data-testid="formula-bar" id="formulaBar" autocomplete="off">
      <button data-testid="save-btn" id="saveBtn">Save</button>
    </div>
    <div id="gridWrap"><table id="grid"></table></div>
    <div data-testid="sheet-tabs" class="tabs" id="sheetTabs"></div>
  </main>
<script>
let workbooks=[], wb=null, activeSheet='Sheet1', activeRef='A1';
const rows=40, cols=20;
const $=id=>document.getElementById(id);
function col(n){let s=''; while(n){let r=(n-1)%26; s=String.fromCharCode(65+r)+s; n=Math.floor((n-1)/26)} return s;}
async function api(path, opts={}){const r=await fetch(path,{headers:{'Content-Type':'application/json'},...opts}); if(!r.ok) throw new Error(await r.text()); return r.json();}
async function loadList(){ const body=await api('/api/workbooks'); workbooks=body.workbooks; if(!workbooks.length){ await api('/api/workbooks',{method:'POST',body:JSON.stringify({name:'Workbook'})}); return loadList(); } renderList(); await loadWorkbook(workbooks[0].id); }
function renderList(){ const list=$('workbookList'); list.innerHTML=''; for(const w of workbooks){ const b=document.createElement('button'); b.className='wb'+(wb&&wb.id===w.id?' active':''); b.dataset.testid='workbook-row-'+w.id; b.textContent=w.name; b.onclick=()=>loadWorkbook(w.id); list.appendChild(b); } }
async function loadWorkbook(id){ wb=await api('/api/workbooks/'+id); activeSheet=wb.sheets[0]?.name||'Sheet1'; renderList(); renderTabs(); renderGrid(); selectCell(activeRef); }
function sheet(){ return wb.sheets.find(s=>s.name===activeSheet) || wb.sheets[0]; }
function cellsMap(){ const m={}; for(const c of (sheet()?.cells||[])) m[c.ref]=c; return m; }
function renderTabs(){ const tabs=$('sheetTabs'); tabs.innerHTML=''; for(const s of wb.sheets){ const b=document.createElement('button'); b.className='tab'+(s.name===activeSheet?' active':''); b.dataset.testid='sheet-tab-'+s.name; b.textContent=s.name; b.onclick=()=>{activeSheet=s.name; renderTabs(); renderGrid();}; tabs.appendChild(b); } }
function renderGrid(){ const m=cellsMap(); const table=$('grid'); table.innerHTML=''; const head=document.createElement('tr'); head.appendChild(Object.assign(document.createElement('th'),{className:'rowhead'})); for(let c=1;c<=cols;c++){ const th=document.createElement('th'); th.textContent=col(c); head.appendChild(th); } table.appendChild(head);
 for(let r=1;r<=rows;r++){ const tr=document.createElement('tr'); const rh=document.createElement('th'); rh.className='rowhead'; rh.textContent=r; tr.appendChild(rh); for(let c=1;c<=cols;c++){ const ref=col(c)+r; const td=document.createElement('td'); td.dataset.testid='cell-'+ref; td.textContent=m[ref]?.display ?? m[ref]?.value ?? ''; td.onclick=()=>selectCell(ref); td.ondblclick=()=>editCell(td,ref); if(ref===activeRef) td.classList.add('active'); tr.appendChild(td);} table.appendChild(tr);} }
function selectCell(ref){ activeRef=ref; $('activeRef').textContent=ref; document.querySelectorAll('td.active').forEach(x=>x.classList.remove('active')); const td=document.querySelector(`[data-testid="cell-${ref}"]`); if(td) td.classList.add('active'); const c=cellsMap()[ref]; $('formulaBar').value=c?.input ?? ''; }
function editCell(td, ref){ const c=cellsMap()[ref]; td.innerHTML=''; const inp=document.createElement('input'); inp.value=c?.input ?? c?.value ?? ''; td.appendChild(inp); inp.focus(); inp.onkeydown=e=>{ if(e.key==='Enter') commit(ref, inp.value); if(e.key==='Escape') renderGrid(); }; inp.onblur=()=>commit(ref, inp.value); }
async function commit(ref, input){ await api(`/api/workbooks/${wb.id}/cells`,{method:'POST',body:JSON.stringify({patches:[{sheet:activeSheet,ref,input}]})}); wb=await api('/api/workbooks/'+wb.id); renderGrid(); selectCell(ref); }
$('formulaBar').addEventListener('keydown', e=>{ if(e.key==='Enter') commit(activeRef, e.target.value); });
$('saveBtn').onclick=async()=>{ wb=await api('/api/workbooks/'+wb.id); renderGrid(); };
$('newBtn').onclick=async()=>{ const n=await api('/api/workbooks',{method:'POST',body:JSON.stringify({name:'Workbook '+(workbooks.length+1)})}); await loadList(); await loadWorkbook(n.id); };
loadList().catch(e=>console.error(e));
</script>
</body>
</html>"""
