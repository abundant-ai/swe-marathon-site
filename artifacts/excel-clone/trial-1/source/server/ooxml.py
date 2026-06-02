"""XLSX import/export via openpyxl."""
from __future__ import annotations
import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

from .refs import col_index_to_letters, parse_a1, normalize_ref, make_a1
from .workbook import Workbook
from .cell import Cell


def import_xlsx(wb: Workbook, data: bytes) -> None:
    """Replace the contents of `wb` with the contents of the .xlsx blob."""
    bio = io.BytesIO(data)
    src = openpyxl.load_workbook(bio, data_only=False)
    # nuke existing sheets
    for s in list(wb.sheets.keys()):
        wb.remove_sheet(s)
    for sn in src.sheetnames:
        sh = wb.add_sheet(sn)
        ws = src[sn]
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                ref = normalize_ref(c.coordinate)
                cell = Cell(sheet=sn, ref=ref)
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    cell.raw_input = v
                elif v is True or v is False:
                    cell.raw_input = "TRUE" if v else "FALSE"
                elif isinstance(v, (int, float)):
                    if isinstance(v, float) and v.is_integer():
                        cell.raw_input = str(int(v))
                    else:
                        cell.raw_input = str(v)
                else:
                    cell.raw_input = str(v)
                # number format
                if c.number_format and c.number_format != "General":
                    cell.format = c.number_format
                # styles
                style = {}
                if c.font:
                    if c.font.bold: style["bold"] = True
                    if c.font.italic: style["italic"] = True
                    if c.font.color and c.font.color.rgb:
                        rgb = c.font.color.rgb
                        if isinstance(rgb, str) and len(rgb) == 8:
                            style["font_color"] = "#" + rgb[2:]
                if c.fill and c.fill.fgColor and c.fill.fgColor.rgb:
                    rgb = c.fill.fgColor.rgb
                    if isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
                        style["bg_color"] = "#" + rgb[2:]
                if style:
                    cell.style = style
                sh.set(cell)
    # named ranges
    try:
        for name, defn in src.defined_names.items():
            try:
                rng = defn.value
            except AttributeError:
                rng = str(defn)
            wb.add_name("workbook", None, name, rng)
    except Exception:
        pass
    wb.recompute_all()


def export_xlsx(wb: Workbook) -> bytes:
    out = openpyxl.Workbook()
    # remove default sheet
    default_name = out.active.title
    if default_name in out.sheetnames:
        del out[default_name]
    for sn in wb.sheet_order:
        ws = out.create_sheet(sn)
        sh = wb.sheets[sn]
        for ref, cell in sh.cells.items():
            # don't export ghosts
            if cell.spill_anchor is not None and (cell.sheet, cell.ref) != cell.spill_anchor:
                continue
            if cell.raw_input is None:
                continue
            c, r, _, _ = parse_a1(ref)
            target = ws.cell(row=r + 1, column=c + 1)
            v = cell.raw_input
            if isinstance(v, str) and v.startswith("="):
                target.value = v
            else:
                # try numeric
                try:
                    if "." in v or "e" in v.lower():
                        target.value = float(v)
                    else:
                        target.value = int(v)
                except (ValueError, TypeError):
                    if v == "TRUE": target.value = True
                    elif v == "FALSE": target.value = False
                    else: target.value = v
            if cell.format:
                target.number_format = cell.format
            if cell.style:
                font_kwargs = {}
                if cell.style.get("bold"):
                    font_kwargs["bold"] = True
                if cell.style.get("italic"):
                    font_kwargs["italic"] = True
                if cell.style.get("font_color"):
                    fc = cell.style["font_color"].lstrip("#").upper()
                    font_kwargs["color"] = "FF" + fc
                if font_kwargs:
                    target.font = Font(**font_kwargs)
                if cell.style.get("bg_color"):
                    bg = cell.style["bg_color"].lstrip("#").upper()
                    target.fill = PatternFill(start_color="FF" + bg, end_color="FF" + bg, fill_type="solid")
    # named ranges
    for n in wb.names.values():
        try:
            d = DefinedName(name=n["name"], attr_text=n["range"])
            out.defined_names[n["name"]] = d
        except Exception:
            pass
    bio = io.BytesIO()
    out.save(bio)
    return bio.getvalue()
